import streamlit as st
import sqlite3
import os
import uuid as uuid_lib
import urllib.parse
from datetime import date, datetime, timedelta
import pandas as pd
import re as re_module
import json
import hashlib
from pathlib import Path
import traceback

# ══════════════════════════════════════════════════════════════════════════════
# إعدادات الصفحة
# ══════════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Orange Lab HVMS",
    page_icon="🟠",
    layout="centered",
    initial_sidebar_state="collapsed",
)

# ══════════════════════════════════════════════════════════════════════════════
# Session State Init
# ══════════════════════════════════════════════════════════════════════════════
for _k, _v in [
    ("authenticated", False), ("user_email", None), ("user_type", None),
    ("page", "home"), ("prefill", {}), ("selected_id", None),
    ("search_q", ""), ("selected_client_phone", ""), ("need_password", False),
    ("current_page", 1), ("page_size", 20), ("total_visits", 0),
]:
    if _k not in st.session_state:
        st.session_state[_k] = _v

# ══════════════════════════════════════════════════════════════════════════════
# Auth Constants
# ══════════════════════════════════════════════════════════════════════════════
ALLOWED_EMAILS = st.secrets.get("allowed_emails", [])
ADMIN_EMAIL   = "Hussein.ali77121@gmail.com"
DIAMOND_EMAIL = "Orangelab511@gmail.com"
LACITE_EMAIL  = "Huossein721@gmail.com"

# ══════════════════════════════════════════════════════════════════════════════
# دوال مساعدة للتنسيق
# ══════════════════════════════════════════════════════════════════════════════
def format_money(value):
    try:
        val = float(value or 0)
        return f"{val:,.0f} جنيه"
    except (ValueError, TypeError):
        return "0 جنيه"

def format_number(value):
    try:
        val = float(value or 0)
        return f"{val:,.0f}"
    except (ValueError, TypeError):
        return "0"

# ══════════════════════════════════════════════════════════════════════════════
# شاشة تسجيل الدخول
# ══════════════════════════════════════════════════════════════════════════════
if not st.session_state.authenticated:
    st.title("🔒 تسجيل الدخول")
    params     = st.query_params
    saved_mail = params.get("remember", "")
    email      = st.text_input("📧 أدخل بريدك الإلكتروني للدخول", value=saved_mail)
    remember_me = st.checkbox("تذكرني في هذا الجهاز", value=bool(saved_mail))
    if st.button("دخول"):
        email_clean = email.strip()
        if email_clean not in ALLOWED_EMAILS:
            st.error("هذا البريد غير مصرح له بالدخول. راجع الأدمن.")
        else:
            if remember_me:
                st.query_params["remember"] = email_clean
            else:
                st.query_params.clear()
            if email_clean.lower() == ADMIN_EMAIL.lower():
                st.session_state.login_email   = email_clean
                st.session_state.need_password = True
                st.rerun()
            else:
                st.session_state.authenticated = True
                st.session_state.user_email    = email_clean
                if email_clean.lower() == DIAMOND_EMAIL.lower():
                    st.session_state.user_type = "diamond"
                elif email_clean.lower() == LACITE_EMAIL.lower():
                    st.session_state.user_type = "lacite"
                else:
                    st.session_state.user_type = "other"
                st.rerun()
    if st.session_state.get("need_password"):
        st.markdown("---")
        st.markdown(f"البريد: **{st.session_state.login_email}**")
        password = st.text_input("🔑 كلمة المرور", type="password")
        if st.button("تأكيد كلمة المرور"):
            correct_password = st.secrets.get("admin_password", "123456")
            if password == correct_password:
                st.success("صلِّ على رسول الله ﷺ - أهلاً بالأدمن")
                st.session_state.authenticated  = True
                st.session_state.user_email     = st.session_state.login_email
                st.session_state.user_type      = "admin"
                st.session_state.need_password  = False
                st.rerun()
            else:
                st.error("كلمة مرور خاطئة")
        if st.button("رجوع"):
            st.session_state.need_password = False
            st.rerun()
    st.markdown("---")
    st.markdown("""
    <div style="text-align:center;margin-top:40px;color:#333;font-size:13px;line-height:1.8;">
      <div style="color:#FF6B00;font-weight:800;font-size:15px;margin-bottom:6px;">📞 للتواصل Contact</div>
      <div><b>Dr / Hussein Ali</b></div>
      <div style="direction:ltr;unicode-bidi:embed;">📱 T: 01016872801</div>
      <div style="direction:ltr;unicode-bidi:embed;">📧 Email: hussein.ali77121@gmail.com</div>
    </div>""", unsafe_allow_html=True)
    st.stop()

# ══════════════════════════════════════════════════════════════════════════════
# إخفاء عناصر Streamlit
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
    <style>
    .stActionButton {display: none !important;}
    #MainMenu {visibility: hidden !important;}
    footer {visibility: hidden !important;}
    header[data-testid="stHeader"] {display: none !important;}
    </style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# قاعدة البيانات + Migrations + الإعدادات الجديدة
# ══════════════════════════════════════════════════════════════════════════════
DB_FILE      = "visits.db"
BACKUP_DIR   = "backups"
BACKUP_EXCEL = "visits_export.xlsx"
SCHEMA_VERSION = 2

@st.cache_resource
def get_connection():
    conn = sqlite3.connect(DB_FILE, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL;")
    return conn

def _ensure_default_doctors(conn):
    count = conn.execute("SELECT COUNT(*) FROM doctors WHERE active=1").fetchone()[0]
    if count == 0:
        DEFAULTS = [
            ("حسين علي", 0), ("ايه جمال", 0), ("محمد شفيق", 1),
            ("شيرين احمد", 0), ("محمد", 0), ("عطيه", 0),
            ("ضي", 0), ("طارق الشافعي", 0),
        ]
        for dname, transport in DEFAULTS:
            try:
                conn.execute(
                    "INSERT OR IGNORE INTO doctors (id,name,transport_eligible,active,created_at) VALUES (?,?,?,1,?)",
                    (uuid_lib.uuid4().hex[:12], dname, transport, datetime.now().isoformat())
                )
            except Exception:
                pass
        conn.commit()

def init_db():
    conn = get_connection()

    # ── الجداول الأساسية ──
    conn.execute("""
        CREATE TABLE IF NOT EXISTS visits (
            id TEXT PRIMARY KEY, created_at TEXT,
            name TEXT NOT NULL, age INTEGER,
            age_unit TEXT DEFAULT 'سنة', phone TEXT NOT NULL,
            visit_date TEXT NOT NULL, visit_time TEXT,
            doctor_name TEXT, branch TEXT DEFAULT 'La Cite',
            address TEXT NOT NULL, location_link TEXT,
            selected_labs_text TEXT, notes TEXT,
            labs_price_before REAL DEFAULT 0, labs_price_after REAL DEFAULT 0,
            transport_fee REAL DEFAULT 0, total_price REAL DEFAULT 0,
            status TEXT DEFAULT 'مجدولة',
            rating INTEGER DEFAULT 0, tag TEXT DEFAULT '',
            payment_status TEXT DEFAULT 'غير مدفوع', payment_method TEXT DEFAULT '',
            paid_amount REAL DEFAULT 0, payment_date TEXT DEFAULT '',
            deleted_at TEXT DEFAULT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS doctors (
            id TEXT PRIMARY KEY, name TEXT NOT NULL,
            phone TEXT DEFAULT '', branch TEXT DEFAULT '',
            commission_pct REAL DEFAULT 5.0,
            transport_eligible INTEGER DEFAULT 0,
            active INTEGER DEFAULT 1, created_at TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS follow_ups (
            id TEXT PRIMARY KEY, visit_id TEXT DEFAULT '',
            client_name TEXT DEFAULT '', client_phone TEXT DEFAULT '',
            follow_up_date TEXT NOT NULL, reason TEXT DEFAULT '',
            done INTEGER DEFAULT 0, created_at TEXT, created_by TEXT DEFAULT ''
        )
    """)

    # ── Audit Log موسع ──
    conn.execute("""
        CREATE TABLE IF NOT EXISTS audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_email TEXT, action TEXT, table_name TEXT,
            record_id TEXT, field_name TEXT DEFAULT '',
            old_value TEXT, new_value TEXT,
            details TEXT, timestamp TEXT
        )
    """)

    # ── Error Log ──
    conn.execute("""
        CREATE TABLE IF NOT EXISTS error_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_email TEXT, page TEXT, exception TEXT,
            stack_trace TEXT, timestamp TEXT
        )
    """)

    # ── Backup Log ──
    conn.execute("""
        CREATE TABLE IF NOT EXISTS backup_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            backup_file TEXT, backup_date TEXT,
            size_bytes INTEGER, status TEXT
        )
    """)

    # ── Settings ──
    conn.execute("""
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY, value TEXT, updated_at TEXT
        )
    """)

    # ── Schema Version ──
    conn.execute("""
        CREATE TABLE IF NOT EXISTS schema_version (
            version INTEGER PRIMARY KEY, applied_at TEXT
        )
    """)

    # ── الفهارس ──
    conn.execute("CREATE INDEX IF NOT EXISTS idx_visits_phone ON visits(phone);")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_visits_visit_date ON visits(visit_date);")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_visits_doctor ON visits(doctor_name);")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_visits_status ON visits(status);")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_visits_payment_status ON visits(payment_status);")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_visits_deleted_at ON visits(deleted_at);")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_visits_branch ON visits(branch);")

    conn.commit()

    # ── إعدادات افتراضية ──
    default_settings = {
        "backup_retention_days": "30",
        "backup_enabled": "true",
        "backup_interval_hours": "24",
    }
    for k, v in default_settings.items():
        conn.execute(
            "INSERT OR IGNORE INTO settings (key, value, updated_at) VALUES (?, ?, ?)",
            (k, v, datetime.now().isoformat())
        )
    conn.commit()

    # ── تسجيل إصدار المخطط ──
    current = conn.execute("SELECT MAX(version) FROM schema_version").fetchone()[0]
    if current is None or current < SCHEMA_VERSION:
        # ── هنا نضع ترقيات المخططات (migrations) ──
        def add_column_if_missing(table, column, definition):
            try:
                conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")
            except sqlite3.OperationalError:
                pass
        columns_to_add = [
            ("visits", "age_unit", "TEXT DEFAULT 'سنة'"),
            ("visits", "status", "TEXT DEFAULT 'مجدولة'"),
            ("visits", "rating", "INTEGER DEFAULT 0"),
            ("visits", "tag", "TEXT DEFAULT ''"),
            ("visits", "payment_status", "TEXT DEFAULT 'غير مدفوع'"),
            ("visits", "payment_method", "TEXT DEFAULT ''"),
            ("visits", "paid_amount", "REAL DEFAULT 0"),
            ("visits", "payment_date", "TEXT DEFAULT ''"),
            ("visits", "deleted_at", "TEXT DEFAULT NULL"),
        ]
        for table, col, defn in columns_to_add:
            add_column_if_missing(table, col, defn)
        conn.execute(
            "INSERT INTO schema_version (version, applied_at) VALUES (?, ?)",
            (SCHEMA_VERSION, datetime.now().isoformat())
        )
        conn.commit()

    _ensure_default_doctors(conn)

init_db()

# ══════════════════════════════════════════════════════════════════════════════
# دوال سجل الأخطاء والنسخ الاحتياطي
# ══════════════════════════════════════════════════════════════════════════════
def log_error(user_email, page, exception, stack_trace=""):
    try:
        conn = get_connection()
        conn.execute("""
            INSERT INTO error_log (user_email, page, exception, stack_trace, timestamp)
            VALUES (?, ?, ?, ?, ?)
        """, (user_email or "", page or "", str(exception), stack_trace, datetime.now().isoformat()))
        conn.commit()
    except Exception:
        pass

def backup_database():
    try:
        if not os.path.exists(BACKUP_DIR):
            os.makedirs(BACKUP_DIR)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_file = os.path.join(BACKUP_DIR, f"visits_backup_{timestamp}.db")
        import shutil
        shutil.copy2(DB_FILE, backup_file)
        conn = get_connection()
        size = os.path.getsize(backup_file)
        conn.execute("""
            INSERT INTO backup_log (backup_file, backup_date, size_bytes, status)
            VALUES (?, ?, ?, ?)
        """, (backup_file, datetime.now().isoformat(), size, "success"))
        conn.commit()
        retention_days = int(get_setting("backup_retention_days", "30"))
        all_backups = sorted([f for f in os.listdir(BACKUP_DIR) if f.startswith("visits_backup_")])
        for f in all_backups[:-retention_days]:
            os.remove(os.path.join(BACKUP_DIR, f))
        return backup_file
    except Exception as e:
        log_error(st.session_state.user_email, "backup", e)
        return None

def restore_database(backup_file):
    try:
        if not os.path.exists(backup_file):
            return False
        import shutil
        shutil.copy2(backup_file, DB_FILE)
        return True
    except Exception as e:
        log_error(st.session_state.user_email, "restore", e)
        return False

def get_setting(key, default=None):
    try:
        conn = get_connection()
        row = conn.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
        return row[0] if row else default
    except Exception:
        return default

def set_setting(key, value):
    try:
        conn = get_connection()
        conn.execute("""
            INSERT OR REPLACE INTO settings (key, value, updated_at)
            VALUES (?, ?, ?)
        """, (key, str(value), datetime.now().isoformat()))
        conn.commit()
    except Exception as e:
        log_error(st.session_state.user_email, "set_setting", e)

# ══════════════════════════════════════════════════════════════════════════════
# ثوابت
# ══════════════════════════════════════════════════════════════════════════════
STATUS_OPTIONS = ["مجدولة", "في الطريق", "تمت", "ملغية"]
STATUS_COLORS  = {"مجدولة":"#3498DB","في الطريق":"#F39C12","تمت":"#27AE60","ملغية":"#E74C3C"}
STATUS_ICONS   = {"مجدولة":"📅","في الطريق":"🚗","تمت":"✅","ملغية":"❌"}
PAYMENT_STATUS_OPTIONS = ["غير مدفوع", "مدفوع جزئياً", "مدفوع"]
PAYMENT_METHODS        = ["نقدي", "محفظة إلكترونية", "تحويل بنكي", "بطاقة"]
PAYMENT_COLORS  = {"غير مدفوع":"#E74C3C","مدفوع جزئياً":"#F39C12","مدفوع":"#27AE60"}
PAYMENT_ICONS   = {"غير مدفوع":"🔴","مدفوع جزئياً":"🟡","مدفوع":"🟢"}
MONTHS_AR = ["يناير","فبراير","مارس","أبريل","مايو","يونيو",
             "يوليو","أغسطس","سبتمبر","أكتوبر","نوفمبر","ديسمبر"]

# ══════════════════════════════════════════════════════════════════════════════
# VISITS CRUD
# ══════════════════════════════════════════════════════════════════════════════
def _clear_tag_cache(phone):
    k = f"_ctag_{phone}"
    if k in st.session_state:
        del st.session_state[k]

def _log_audit(user_email, action, table_name, record_id, field_name="", old_value="", new_value="", details=""):
    try:
        conn = get_connection()
        conn.execute("""
            INSERT INTO audit_log (user_email, action, table_name, record_id, field_name,
                                    old_value, new_value, details, timestamp)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (user_email or "", action, table_name, record_id, field_name,
              str(old_value), str(new_value), details, datetime.now().isoformat()))
        conn.commit()
    except Exception as e:
        log_error(user_email, "audit_log", e)

def fetch_visit_by_unique_keys(name, phone, visit_date):
    conn = get_connection()
    row = conn.execute(
        "SELECT * FROM visits WHERE name=? AND phone=? AND visit_date=? AND deleted_at IS NULL LIMIT 1",
        (name, phone, visit_date)
    ).fetchone()
    return dict(row) if row else None

def fetch_visits(filters=None, page=None, page_size=None):
    conn = get_connection()
    query = "SELECT * FROM visits WHERE deleted_at IS NULL"
    params = []
    if filters:
        if filters.get("search"):
            s = f"%{filters['search']}%"
            query += " AND (name LIKE ? OR phone LIKE ? OR address LIKE ? OR doctor_name LIKE ? OR id LIKE ?)"
            params.extend([s, s, s, s, s])
        if filters.get("branch"):
            query += " AND branch = ?"
            params.append(filters["branch"])
        if filters.get("doctor"):
            query += " AND doctor_name = ?"
            params.append(filters["doctor"])
        if filters.get("month") and filters.get("year"):
            y, m = filters["year"], filters["month"]
            query += " AND strftime('%Y',visit_date)=? AND strftime('%m',visit_date)=?"
            params.extend([str(y), f"{m:02d}"])
        if filters.get("date_exact"):
            query += " AND visit_date = ?"
            params.append(filters["date_exact"])
        if filters.get("date_from"):
            query += " AND visit_date >= ?"
            params.append(filters["date_from"])
        if filters.get("date_to"):
            query += " AND visit_date <= ?"
            params.append(filters["date_to"])
        if filters.get("status"):
            query += " AND status = ?"
            params.append(filters["status"])
        if filters.get("payment_status"):
            query += " AND payment_status = ?"
            params.append(filters["payment_status"])
    query += " ORDER BY visit_date ASC, visit_time ASC"

    count_query = f"SELECT COUNT(*) FROM ({query})"
    total = conn.execute(count_query, params).fetchone()[0]

    if page is not None and page_size is not None:
        offset = (page - 1) * page_size
        query += f" LIMIT {page_size} OFFSET {offset}"

    rows = conn.execute(query, params).fetchall()
    visits = [dict(r) for r in rows]
    return visits, total

def fetch_visit_by_id(vid):
    conn = get_connection()
    row = conn.execute("SELECT * FROM visits WHERE id=? AND deleted_at IS NULL",(vid,)).fetchone()
    return dict(row) if row else None

def fetch_client_history(phone, exclude_id=None, limit=None):
    conn = get_connection()
    if exclude_id:
        rows = conn.execute(
            "SELECT * FROM visits WHERE phone=? AND id!=? AND deleted_at IS NULL ORDER BY visit_date DESC",
            (phone, exclude_id)
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM visits WHERE phone=? AND deleted_at IS NULL ORDER BY visit_date DESC",(phone,)
        ).fetchall()
    if limit and len(rows) > limit:
        rows = rows[:limit]
    return [dict(r) for r in rows]

def insert_visit(record):
    conn = get_connection()
    conn.execute("""
        INSERT INTO visits (
            id,created_at,name,age,age_unit,phone,visit_date,visit_time,
            doctor_name,branch,address,location_link,selected_labs_text,notes,
            labs_price_before,labs_price_after,transport_fee,total_price,status,
            payment_status,payment_method,paid_amount,payment_date
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        record["id"], record["created_at"], record["name"], record["age"],
        record.get("age_unit","سنة"), record["phone"],
        record["visit_date"], record["visit_time"], record["doctor_name"],
        record.get("branch","La Cite"), record["address"], record["location_link"],
        record["selected_labs_text"], record["notes"],
        record["labs_price_before"], record["labs_price_after"],
        record["transport_fee"], record["total_price"],
        record.get("status","مجدولة"),
        record.get("payment_status","غير مدفوع"),
        record.get("payment_method",""), record.get("paid_amount",0),
        record.get("payment_date",""),
    ))
    conn.commit()
    _clear_tag_cache(record["phone"])
    _log_audit(record.get("_user",""), "insert", "visits", record["id"], details=f"اسم: {record['name']}")

def update_visit(record):
    conn = get_connection()
    old = fetch_visit_by_id(record["id"])
    conn.execute("""
        UPDATE visits SET
            name=?,age=?,age_unit=?,phone=?,visit_date=?,visit_time=?,
            doctor_name=?,branch=?,address=?,location_link=?,
            selected_labs_text=?,notes=?,labs_price_before=?,
            labs_price_after=?,transport_fee=?,total_price=?,status=?,
            payment_status=?,payment_method=?,paid_amount=?,payment_date=?
        WHERE id=?
    """, (
        record["name"], record["age"], record.get("age_unit","سنة"),
        record["phone"], record["visit_date"], record["visit_time"],
        record["doctor_name"], record.get("branch","La Cite"),
        record["address"], record["location_link"], record["selected_labs_text"],
        record["notes"], record["labs_price_before"], record["labs_price_after"],
        record["transport_fee"], record["total_price"],
        record.get("status","مجدولة"),
        record.get("payment_status","غير مدفوع"),
        record.get("payment_method",""), record.get("paid_amount",0),
        record.get("payment_date",""), record["id"]
    ))
    conn.commit()
    _clear_tag_cache(record["phone"])
    if old:
        for field in ["name","age","phone","visit_date","doctor_name","branch","address","labs_price_before","labs_price_after","transport_fee","total_price","status","payment_status"]:
            if old.get(field) != record.get(field):
                _log_audit(record.get("_user",""), "update", "visits", record["id"],
                           field_name=field, old_value=old.get(field), new_value=record.get(field))

def soft_delete_visit(vid, user_email=""):
    conn = get_connection()
    conn.execute("UPDATE visits SET deleted_at=? WHERE id=?",(datetime.now().isoformat(), vid))
    conn.commit()
    _log_audit(user_email, "soft_delete", "visits", vid)

def update_status_only(vid, new_status):
    conn = get_connection()
    old = fetch_visit_by_id(vid)
    conn.execute("UPDATE visits SET status=? WHERE id=?",(new_status, vid))
    conn.commit()
    if old:
        _log_audit(st.session_state.user_email, "update_status", "visits", vid, field_name="status", old_value=old.get("status"), new_value=new_status)

def update_rating(vid, rating):
    conn = get_connection()
    old = fetch_visit_by_id(vid)
    conn.execute("UPDATE visits SET rating=? WHERE id=?",(rating, vid))
    conn.commit()
    if old:
        _log_audit(st.session_state.user_email, "update_rating", "visits", vid, field_name="rating", old_value=old.get("rating"), new_value=rating)

def update_tag(vid, tag):
    conn = get_connection()
    old = fetch_visit_by_id(vid)
    conn.execute("UPDATE visits SET tag=? WHERE id=?",(tag, vid))
    conn.commit()
    if old:
        _log_audit(st.session_state.user_email, "update_tag", "visits", vid, field_name="tag", old_value=old.get("tag"), new_value=tag)

def update_payment(vid, pay_status, pay_method, paid_amount, pay_date, user_email=""):
    conn = get_connection()
    old = fetch_visit_by_id(vid)
    conn.execute(
        "UPDATE visits SET payment_status=?,payment_method=?,paid_amount=?,payment_date=? WHERE id=?",
        (pay_status, pay_method, float(paid_amount), pay_date, vid)
    )
    conn.commit()
    if old:
        _log_audit(user_email, "update_payment", "visits", vid, field_name="payment_status", old_value=old.get("payment_status"), new_value=pay_status)

# ══════════════════════════════════════════════════════════════════════════════
# DOCTORS CRUD
# ══════════════════════════════════════════════════════════════════════════════
def fetch_doctors(active_only=True):
    conn = get_connection()
    q = "SELECT * FROM doctors" + (" WHERE active=1" if active_only else "")
    q += " ORDER BY name ASC"
    return [dict(r) for r in conn.execute(q).fetchall()]

def get_doctor_names(active_only=True):
    names = [d["name"] for d in fetch_doctors(active_only)]
    return names if names else ["حسين علي","ايه جمال","محمد شفيق","شيرين احمد","محمد","عطيه","ضي","طارق الشافعي"]

def get_transport_eligible_doctors():
    conn = get_connection()
    rows = conn.execute("SELECT name FROM doctors WHERE transport_eligible=1 AND active=1").fetchall()
    return [r[0] for r in rows] or ["محمد شفيق"]

def insert_doctor(name, commission_pct=5.0, transport_eligible=0, phone="", branch=""):
    conn = get_connection()
    doc_id = uuid_lib.uuid4().hex[:12]
    conn.execute(
        "INSERT INTO doctors (id,name,phone,branch,commission_pct,transport_eligible,active,created_at) VALUES (?,?,?,?,?,?,1,?)",
        (doc_id, name, phone, branch, commission_pct, transport_eligible, datetime.now().isoformat())
    )
    conn.commit()
    _log_audit(st.session_state.user_email, "insert", "doctors", doc_id, details=name)
    return doc_id

def update_doctor(doc_id, name, commission_pct, transport_eligible, phone, branch):
    conn = get_connection()
    conn.execute(
        "UPDATE doctors SET name=?,commission_pct=?,transport_eligible=?,phone=?,branch=? WHERE id=?",
        (name, commission_pct, transport_eligible, phone, branch, doc_id)
    )
    conn.commit()
    _log_audit(st.session_state.user_email, "update", "doctors", doc_id)

def toggle_doctor_active(doc_id):
    conn = get_connection()
    conn.execute("UPDATE doctors SET active=1-active WHERE id=?",(doc_id,))
    conn.commit()
    _log_audit(st.session_state.user_email, "toggle_active", "doctors", doc_id)

# ══════════════════════════════════════════════════════════════════════════════
# FOLLOW-UPS CRUD
# ══════════════════════════════════════════════════════════════════════════════
def fetch_follow_ups(filters=None):
    conn = get_connection()
    query = "SELECT * FROM follow_ups"
    conditions, params = [], []
    if filters:
        if filters.get("done") is not None:
            conditions.append("done=?"); params.append(int(filters["done"]))
        if filters.get("date_from"):
            conditions.append("follow_up_date>=?"); params.append(filters["date_from"])
        if filters.get("date_to"):
            conditions.append("follow_up_date<=?"); params.append(filters["date_to"])
        if filters.get("phone"):
            conditions.append("client_phone=?"); params.append(filters["phone"])
    if conditions:
        query += " WHERE " + " AND ".join(conditions)
    query += " ORDER BY follow_up_date ASC"
    return [dict(r) for r in conn.execute(query, params).fetchall()]

def insert_follow_up(visit_id, client_name, client_phone, follow_up_date, reason, created_by=""):
    conn = get_connection()
    fu_id = uuid_lib.uuid4().hex[:12]
    conn.execute("""
        INSERT INTO follow_ups (id,visit_id,client_name,client_phone,follow_up_date,reason,done,created_at,created_by)
        VALUES (?,?,?,?,?,?,0,?,?)
    """, (fu_id, visit_id, client_name, client_phone, follow_up_date, reason, datetime.now().isoformat(), created_by))
    conn.commit()
    _log_audit(created_by, "insert", "follow_ups", fu_id, details=f"لـ {client_name}")
    return fu_id

def complete_follow_up(fu_id):
    conn = get_connection()
    conn.execute("UPDATE follow_ups SET done=1 WHERE id=?",(fu_id,))
    conn.commit()

def delete_follow_up(fu_id):
    conn = get_connection()
    conn.execute("DELETE FROM follow_ups WHERE id=?",(fu_id,))
    conn.commit()

# ══════════════════════════════════════════════════════════════════════════════
# Excel Export/Import
# ══════════════════════════════════════════════════════════════════════════════
def export_to_excel(branch_filter=None, month=None, year=None, date_from=None, date_to=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    filters = {}
    if branch_filter:
        filters["branch"] = branch_filter
    if date_from and date_to:
        filters["date_from"] = str(date_from); filters["date_to"] = str(date_to)
    elif month and year:
        filters["month"] = month; filters["year"] = year
    visits, _ = fetch_visits(filters if filters else None, page=None, page_size=None)
    df = pd.DataFrame(visits)
    ORANGE = "FF6B00"; ORANGE_LIGHT = "FFF3E8"; DOC_HDR_BG = "2C3E50"; WHITE = "FFFFFF"
    STATUS_FILL = {"تمت":"D5F5E3","ملغية":"FADBD8","في الطريق":"FEF9E7","مجدولة":"D6EAF8"}
    thin = Side(style="thin", color="FFBB80")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    RIGHT  = Alignment(horizontal="right",  vertical="center", wrap_text=True)
    MAIN_COLS = [
        ("status","الحالة"),("total_price","الإجمالي"),("transport_fee","بدل الانتقال"),
        ("labs_price_after","السعر بعد الخصم"),("labs_price_before","السعر قبل الخصم"),
        ("selected_labs_text","التحاليل"),("address","العنوان"),("doctor_name","الدكتور"),
        ("visit_date","تاريخ الزيارة"),("phone","التليفون"),("name","الاسم"),
    ]
    PRICE_KEYS = {"labs_price_before","labs_price_after","transport_fee","total_price"}
    col_keys = [c[0] for c in MAIN_COLS]; col_labels = [c[1] for c in MAIN_COLS]
    n_cols = len(MAIN_COLS)
    WIDTHS_MAIN = {"status":10,"total_price":14,"transport_fee":14,"labs_price_after":16,
                   "labs_price_before":16,"selected_labs_text":28,"address":32,
                   "doctor_name":14,"visit_date":13,"phone":15,"name":20}
    if date_from and date_to:
        month_label = f"{date_from} → {date_to}"
        period = f"{date_from}_to_{date_to}"
    elif month and year:
        month_label = f"{MONTHS_AR[month-1]} {year}"
        period = f"{MONTHS_AR[month-1]}_{year}"
    else:
        month_label = str(date.today()); period = str(date.today())
    branch_label = f"فرع {branch_filter}" if branch_filter else "كل الفروع"
    if branch_filter == "Diamond":    fname = f"diamond_{period}.xlsx"
    elif branch_filter == "La Cite":  fname = f"lacite_{period}.xlsx"
    else: fname = f"visits_{period}.xlsx" if (date_from or month) else BACKUP_EXCEL
    if df.empty:
        pd.DataFrame().to_excel(fname, index=False, engine="openpyxl"); return df, fname
    df_valid = df[df["status"] != "ملغية"]; n_rows = len(df)
    wb = Workbook(); ws = wb.active; ws.title = "الزيارات"; ws.sheet_view.rightToLeft = True
    DATA_START = 4; DATA_END = DATA_START + n_rows - 1; TOTAL_ROW = DATA_END + 1
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=n_cols)
    c = ws.cell(1,1,f"🟠 Orange Lab HVMS — {branch_label} — {month_label}")
    c.font = Font(name="Cairo",bold=True,color=WHITE,size=14)
    c.fill = PatternFill("solid",fgColor=ORANGE); c.alignment = CENTER; ws.row_dimensions[1].height = 38
    sl = ["تمت","مجدولة","في الطريق","ملغية"]
    parts = "  |  ".join(f"{s}: {(df['status']==s).sum()}" for s in sl if (df['status']==s).sum()>0)
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=n_cols)
    c = ws.cell(2,1,f"إجمالي: {n_rows} زيارة  |  {parts}  |  تاريخ التصدير: {date.today()}")
    c.font = Font(name="Cairo",bold=True,color=ORANGE,size=10)
    c.fill = PatternFill("solid",fgColor=ORANGE_LIGHT); c.alignment = CENTER; ws.row_dimensions[2].height = 22
    for ci, label in enumerate(col_labels, 1):
        c = ws.cell(3,ci,label)
        c.font = Font(name="Cairo",bold=True,color=WHITE,size=11)
        c.fill = PatternFill("solid",fgColor=ORANGE); c.alignment = CENTER; c.border = BORDER
    ws.row_dimensions[3].height = 30
    for ri, (_, row) in enumerate(df.iterrows(), start=DATA_START):
        st_val = str(row.get("status",""))
        fc = STATUS_FILL.get(st_val, WHITE)
        for ci, key in enumerate(col_keys, 1):
            val = row.get(key, "")
            if pd.isna(val): val = ""
            c = ws.cell(ri,ci,val)
            c.font = Font(name="Cairo",size=10); c.fill = PatternFill("solid",fgColor=fc); c.border = BORDER
            if key in PRICE_KEYS:
                c.number_format = '#,##0 "ج"'; c.alignment = Alignment(horizontal="center",vertical="center")
            elif key == "selected_labs_text":
                c.alignment = Alignment(horizontal="right",vertical="top",wrap_text=True)
            else: c.alignment = RIGHT
        ws.row_dimensions[ri].height = 22
    labs_ci = col_keys.index("selected_labs_text") + 1
    ws.merge_cells(start_row=TOTAL_ROW,start_column=labs_ci,end_row=TOTAL_ROW,end_column=n_cols)
    c = ws.cell(TOTAL_ROW,labs_ci,"الإجمالي الكلي")
    c.font = Font(name="Cairo",bold=True,color=WHITE,size=12)
    c.fill = PatternFill("solid",fgColor=ORANGE); c.alignment = CENTER; c.border = BORDER
    for ci, key in enumerate(col_keys, 1):
        if key in PRICE_KEYS:
            cl = get_column_letter(ci)
            c = ws.cell(TOTAL_ROW,ci,f"=SUM({cl}{DATA_START}:{cl}{DATA_END})")
            c.font = Font(name="Cairo",bold=True,color=WHITE,size=12)
            c.fill = PatternFill("solid",fgColor=ORANGE); c.alignment = Alignment(horizontal="center",vertical="center")
            c.border = BORDER; c.number_format = '#,##0 "ج"'
        elif key == "status":
            c = ws.cell(TOTAL_ROW,ci,f"{n_rows} زيارة")
            c.font = Font(name="Cairo",bold=True,color=WHITE,size=12)
            c.fill = PatternFill("solid",fgColor=ORANGE); c.alignment = CENTER; c.border = BORDER
    ws.row_dimensions[TOTAL_ROW].height = 28
    DOC_COLS = ["م","اسم الدكتور","عدد الزيارات","إجمالي بعد الخصم","بدل الزيارة (5%)","بدل الانتقال","الإجمالي"]
    n_doc_cols = len(DOC_COLS); DOC_TITLE = TOTAL_ROW+2; DOC_HDR = TOTAL_ROW+3; DOC_DATA = TOTAL_ROW+4
    ws.row_dimensions[TOTAL_ROW+1].height = 16
    ws.merge_cells(start_row=DOC_TITLE,start_column=1,end_row=DOC_TITLE,end_column=n_doc_cols)
    c = ws.cell(DOC_TITLE,1,"📊 ملخص الأطباء — بدل الزيارات")
    c.font = Font(name="Cairo",bold=True,color=WHITE,size=12)
    c.fill = PatternFill("solid",fgColor=DOC_HDR_BG); c.alignment = CENTER; ws.row_dimensions[DOC_TITLE].height = 26
    for ci, label in enumerate(DOC_COLS, 1):
        c = ws.cell(DOC_HDR,ci,label)
        c.font = Font(name="Cairo",bold=True,color=WHITE,size=11)
        c.fill = PatternFill("solid",fgColor=DOC_HDR_BG); c.alignment = CENTER; c.border = BORDER
    ws.row_dimensions[DOC_HDR].height = 28
    TRANSPORT_ELIGIBLE = get_transport_eligible_doctors()
    doc_data = {}
    for _, row in df_valid.iterrows():
        doc = str(row.get("doctor_name","غير محدد") or "غير محدد")
        if doc not in doc_data: doc_data[doc] = {"count":0,"after":0,"transport":0}
        doc_data[doc]["count"] += 1
        doc_data[doc]["after"] += float(row.get("labs_price_after",0) or 0)
        if doc in TRANSPORT_ELIGIBLE:
            doc_data[doc]["transport"] += float(row.get("transport_fee",0) or 0)
    docs_sorted = sorted(doc_data.items(), key=lambda x: x[1]["count"], reverse=True)
    for idx, (doc_name, d) in enumerate(docs_sorted):
        ri = DOC_DATA + idx
        allowance = d["after"] * 0.05; total_doc = allowance + d["transport"]
        fc = WHITE if idx % 2 == 0 else "F5F5F5"
        for ci, val in enumerate([idx+1, doc_name, d["count"], d["after"], allowance, d["transport"], total_doc], 1):
            c = ws.cell(ri,ci,val)
            c.font = Font(name="Cairo",size=11); c.fill = PatternFill("solid",fgColor=fc)
            c.border = BORDER; c.alignment = CENTER if ci != 2 else RIGHT
            if ci >= 4: c.number_format = '#,##0.## "ج"'
        ws.row_dimensions[ri].height = 22
    n_docs = len(docs_sorted); DOC_TOTAL_ROW = DOC_DATA + n_docs
    tot_after = sum(d["after"] for _,d in docs_sorted)
    tot_allow = tot_after * 0.05
    tot_trans = sum(d["transport"] for _,d in docs_sorted)
    tot_grand = tot_allow + tot_trans
    ws.merge_cells(start_row=DOC_TOTAL_ROW,start_column=1,end_row=DOC_TOTAL_ROW,end_column=2)
    c = ws.cell(DOC_TOTAL_ROW,1,"الإجمالي")
    c.font = Font(name="Cairo",bold=True,color=WHITE,size=12)
    c.fill = PatternFill("solid",fgColor=ORANGE); c.alignment = CENTER; c.border = BORDER
    for ci, val in [(3,sum(d["count"] for _,d in docs_sorted)),(4,tot_after),(5,tot_allow),(6,tot_trans),(7,tot_grand)]:
        c = ws.cell(DOC_TOTAL_ROW,ci,val)
        c.font = Font(name="Cairo",bold=True,color=WHITE,size=12)
        c.fill = PatternFill("solid",fgColor=ORANGE)
        c.alignment = Alignment(horizontal="center",vertical="center"); c.border = BORDER
        if ci >= 4: c.number_format = '#,##0.## "ج"'
    ws.row_dimensions[DOC_TOTAL_ROW].height = 26
    for ci, key in enumerate(col_keys, 1):
        ws.column_dimensions[get_column_letter(ci)].width = WIDTHS_MAIN.get(key, 14)
    for ci, w in enumerate([5,18,14,18,16,14,14], 1):
        cur = ws.column_dimensions[get_column_letter(ci)].width
        ws.column_dimensions[get_column_letter(ci)].width = max(cur, w)
    ws.freeze_panes = "A4"; wb.save(fname)
    return df, fname

# ══════════════════════════════════════════════════════════════════════════════
# 🔥 دالة الاستيراد الذكية الجديدة (نسخة مقاومة للأخطاء) 🔥
# ══════════════════════════════════════════════════════════════════════════════
def import_from_excel(uploaded_file):
    try:
        # قراءة ملف الإكسيل (الورقة الأولى دائماً)
        # قراءة ذكية لملفات Excel (ملفات البرنامج أو ملفات عادية)

        headers_to_try = [2, 3, 0]
        df = None

        for hdr in headers_to_try:
           try:
        uploaded_file.seek(0)
        tmp = pd.read_excel(
            uploaded_file,
            engine="openpyxl",
            sheet_name=0,
            header=hdr
        )

        cols = [str(c).strip().replace(" ", "").replace("\xa0", "") for c in tmp.columns]

        if any(c in cols for c in ["الاسم", "name"]):
            df = tmp
            break

    except Exception:
        pass

if df is None:
    uploaded_file.seek(0)
    df = pd.read_excel(uploaded_file, engine="openpyxl", sheet_name=0)

# حذف الصفوف الفارغة
df = df.dropna(how="all")
        
        # تنظيف عنيف جداً لأسماء الأعمدة (حتى من المسافات المخفية)
        cleaned_cols = []
        for c in df.columns:
            c = str(c).strip().replace(" ", "").replace("\xa0", "") # إزالة المسافات العادية والمخفية
            cleaned_cols.append(c)
        df.columns = cleaned_cols

        count_imported = 0
        count_updated = 0

        # خريطة متطابقة مع كل أنواع الملفات (بما فيها الملفات اللي بعتها)
        col_mapping = {}
        name_mapping = {
            "id": ["رقمالزيارة", "id", "visitid", "معرفالزيارة", "زيارةرقم"],
            "name": ["الاسم", "اسمالمريض", "اسم", "name", "patientname"],
            "phone": ["التليفون", "رقمالتليفون", "هاتف", "phone", "mobile", "telephone"],
            "visit_date": ["تاريخالزيارة", "تاريخ", "الزيارة", "visitdate", "visit_date", "التاريخ"],
            "age": ["السن", "العمر", "age"],
            "age_unit": ["الوحدة", "وحدةالعمر", "العمربوحدة", "age_unit"],
            "doctor_name": ["الدكتور", "اسمالدكتور", "طبيب", "doctor"],
            "branch": ["الفرع", "branch"],
            "address": ["العنوان", "address"],
            "location_link": ["رابطالموقع", "location", "map", "رابط", "الموقع"],
            "selected_labs_text": ["التحاليل", "التحليل", "labs", "tests", "selected_labs_text"],
            "notes": ["ملاحظات", "notes"],
            "labs_price_before": ["السعرقبلالخصم", "قبلالخصم", "pricebefore", "labs_price_before", "الخصم"],
            "labs_price_after": ["السعربعدالخصم", "بعدالخصم", "priceafter", "labs_price_after"],
            "transport_fee": ["بدلالانتقال", "النقل", "transportfee", "transport_fee", "الانتقال"],
            "total_price": ["الإجمالي", "السعرالاجمالي", "الاجمالي", "total"],
            "status": ["الحالة", "status", "الوضع"],
            "payment_status": ["حالةالدفع", "paymentstatus", "دفع", "payment_status"],
            "payment_method": ["طريقةالدفع", "paymentmethod", "payment_method"],
            "paid_amount": ["المبلغالمدفوع", "paidamount", "مدفوع", "paid"],
            "visit_time": ["الوقت", "الموعد", "visit_time", "time"],
            "created_at": ["تاريخالإنشاء", "createdat", "created_at", "تاريخانشاء"]
        }

        # بناء خريطة الأعمدة
        for internal_key, possible_names in name_mapping.items():
            for col in df.columns:
                if col.lower() in [n.lower() for n in possible_names]:
                    col_mapping[internal_key] = col
                    break

        # (للتصحيح فقط) عرض الأعمدة اللي التطبيق قدر يتعرف عليها في الشريط الجانبي
        st.sidebar.write("✅ الأعمدة التي تم التعرف عليها في الملف:", col_mapping)

        # 2. المرور على كل صف
        for _, row in df.iterrows():
            record = {}
            
            # تعيين القيم من ملف الإكسيل
            for key in col_mapping:
                val = row[col_mapping[key]]
                if pd.isna(val): val = None
                record[key] = val

            # معالجة الأرقام (ضمان إنها أرقام وليست نصوص)
            for num_key in ["labs_price_before", "labs_price_after", "transport_fee", "total_price", "paid_amount", "age"]:
                if num_key in record and record[num_key] is not None:
                    try:
                        if isinstance(record[num_key], str):
                            record[num_key] = float(record[num_key].replace(",", ""))
                        else:
                            record[num_key] = float(record[num_key])
                    except:
                        record[num_key] = 0
                elif num_key not in record:
                    record[num_key] = 0

            # معالجة القيم النصية الافتراضية
            for txt_key in ["status", "branch", "payment_status", "payment_method", "doctor_name", "age_unit"]:
                if txt_key not in record or record[txt_key] is None:
                    if txt_key == "status": record[txt_key] = "مجدولة"
                    elif txt_key == "branch": record[txt_key] = "La Cite"
                    elif txt_key == "payment_status": record[txt_key] = "غير مدفوع"
                    elif txt_key == "payment_method": record[txt_key] = "نقدي"
                    elif txt_key == "age_unit": record[txt_key] = "سنة"
                    else: record[txt_key] = ""

            # معالجة التواريخ بدقة
            if "visit_date" in record and record["visit_date"] is not None:
                try:
                    record["visit_date"] = pd.to_datetime(record["visit_date"]).strftime("%Y-%m-%d")
                except:
                    record["visit_date"] = date.today().isoformat()
            else:
                record["visit_date"] = date.today().isoformat()
                
            # إذا لم يكن هناك وقت، ضع فارغاً
            if "visit_time" not in record or record["visit_time"] is None:
                record["visit_time"] = ""

            # حساب الإجمالي إذا لم يكن موجوداً
            if record.get("total_price", 0) == 0:
                record["total_price"] = record.get("labs_price_after", 0) + record.get("transport_fee", 0)

            # التأكد من وجود البيانات الأساسية
            if not record.get("name") or not record.get("phone"):
                continue

            # إنشاء ID إذا لم يكن موجوداً في الملف
            if "id" not in record or not record["id"]:
                record["id"] = uuid_lib.uuid4().hex[:16]
            else:
                record["id"] = str(record["id"])

            # تحديد ما إذا كان السجل موجوداً (بناءً على الـ ID أو بناءً على الاسم والتليفون والتاريخ)
            existing_record = None
            
            # 1. حاول البحث بالـ ID
            if "id" in record and record["id"]:
                existing_record = fetch_visit_by_id(record["id"])
            
            # 2. إذا لم يوجد بالـ ID، ابحث بالاسم + التليفون + التاريخ (لمنع التكرار للملفات التي ليس بها ID)
            if not existing_record and record.get("name") and record.get("phone") and record.get("visit_date"):
                existing_record = fetch_visit_by_unique_keys(
                    record.get("name"), 
                    record.get("phone"), 
                    record.get("visit_date")
                )

            # 3. إما التحديث أو الإدراج
            if existing_record:
                record["id"] = existing_record["id"] # نستخدم الـ ID الموجود لتحديث السجل بشكل صحيح
                record["_user"] = st.session_state.get("user_email", "system_import")
                update_visit(record)
                count_updated += 1
            else:
                record["created_at"] = datetime.now().isoformat()
                record["_user"] = st.session_state.get("user_email", "system_import")
                insert_visit(record)
                count_imported += 1

        return count_imported, count_updated

    except Exception as e:
        st.error(f"حدث خطأ أثناء معالجة الملف: {e}")
        return 0, 0

# ══════════════════════════════════════════════════════════════════════════════
# Quick Panels واقتراحات (كما هي)
# ══════════════════════════════════════════════════════════════════════════════
QUICK_PANELS = [
    {"name":"🩸 CBC",      "tests":["CBC"]},
    {"name":"🍬 Diabetes", "tests":["HbA1C","Urea","Creatinine (Serum)","Uric Acid","ALT (SGPT)","AST (SGOT)","Urine Examination"]},
    {"name":"❤️ Cardiac",  "tests":["Cholesterol","HDL","LDL","Triglycerides","ALT (SGPT)","AST (SGOT)","Uric Acid"]},
    {"name":"🦋 Thyroid",  "tests":["TSH","FT3","FT4"]},
    {"name":"🔋 Fatigue",  "tests":["CBC","Ferritin","Vitamin D3(25 Hydroxy Cholecal.)","TSH"]},
    {"name":"🧪 Kidney",   "tests":["Urea","Creatinine (Serum)","Uric Acid","Urine Examination"]},
    {"name":"🫀 Liver",    "tests":["ALT (SGPT)","AST (SGOT)","Albumin (ALB)","Bilirubin Total","Alkaline Phosphatase (ALP)"]},
    {"name":"🌟 General",  "tests":["CBC","Cholesterol","HDL","LDL","Triglycerides","HbA1C","TSH","ALT (SGPT)","AST (SGOT)","Urea","Creatinine (Serum)","Urine Examination"]},
]

BUNDLE_RULES = [
    {"trigger":["Vitamin D3(25 Hydroxy Cholecal.)"],
     "bundle":"Vitamin D3 Couple","note":"💡 عرض الفردين بسعر أوفر — وفّر فلوس!",
     "saving_fn": lambda pl: (pl.get("Vitamin D3(25 Hydroxy Cholecal.)",400)*2 - pl.get("Vitamin D3 Couple",pl.get("Vitamin D3(25 Hydroxy Cholecal.)",400)+200)),
     "fallback_add":"Vitamin D3(25 Hydroxy Cholecal.)","fallback_note":"💡 إضافة فيتامين د ثاني للفردين بخصم خاص"},
]

PANEL_SUGGEST_RULES = [
    {"panel":"🍬 Diabetes","core":["HbA1C"],"suggest":["Urea","Creatinine (Serum)","ALT (SGPT)","Urine Examination","Uric Acid","AST (SGOT)"],"reason":"متابعة مريض السكر — يُنصح بقياس وظائف الكلى والكبد"},
    {"panel":"❤️ Cardiac","core":["Cholesterol","Triglycerides"],"suggest":["HDL","LDL","ALT (SGPT)","Uric Acid"],"reason":"تقييم القلب والأوعية الدموية الكامل"},
    {"panel":"🦋 Thyroid","core":["TSH"],"suggest":["FT3","FT4"],"reason":"TSH وحده غير كافٍ — يُنصح بالثلاثي الكامل"},
    {"panel":"🧪 Kidney","core":["Urea","Creatinine (Serum)"],"suggest":["Uric Acid","Urine Examination"],"reason":"تقييم وظائف الكلى الكامل"},
    {"panel":"🫀 Liver","core":["ALT (SGPT)","AST (SGOT)"],"suggest":["Albumin (ALB)","Bilirubin Total","Alkaline Phosphatase (ALP)","GGT"],"reason":"تقييم وظائف الكبد الكامل"},
    {"panel":"🔋 Fatigue","core":["CBC","Ferritin"],"suggest":["Vitamin D3(25 Hydroxy Cholecal.)","TSH","B12"],"reason":"الإجهاد المزمن — شيّك على الفيتامينات والغدة"},
    {"panel":"🩸 Anemia","core":["CBC"],"suggest":["Ferritin","Iron (Serum)","TIBC","B12","Folic Acid"],"reason":"CBC وحده لا يكفي لتشخيص الأنيميا — أضف مؤشرات الحديد"},
    {"panel":"🌟 General","core":["CBC","Cholesterol","HbA1C"],"suggest":["TSH","Vitamin D3(25 Hydroxy Cholecal.)","Ferritin"],"reason":"فحص شامل — أضف فيتامين د والغدة الدرقية"},
]

CLINICAL_RULES = [
    {"if_present":["HbA1C","Cholesterol"],"suggest":["Creatinine (Serum)","ALT (SGPT)"],"reason":"مريض سكر + دهون → وظائف كلى وكبد ضرورية"},
    {"if_present":["TSH"],"suggest":["Cholesterol","CBC"],"reason":"اضطراب الغدة الدرقية يؤثر على الدهون والدم"},
    {"if_present":["Ferritin"],"suggest":["CBC","Iron (Serum)"],"reason":"فيريتين بدون صورة دم ومخزون حديد غير مكتمل"},
    {"if_present":["Creatinine (Serum)","Urea"],"suggest":["Urine Examination","Uric Acid"],"reason":"وظائف كلى — أضف تحليل بول وحمض يوريك"},
    {"if_present":["ALT (SGPT)","AST (SGOT)"],"suggest":["Albumin (ALB)","Bilirubin Total"],"reason":"وظائف كبد جزئية — أضف الألبيومين والبيليروبين"},
    {"if_present":["Vitamin D3(25 Hydroxy Cholecal.)"],"suggest":["Calcium (Serum)","Phosphorus (Serum)"],"reason":"فيتامين د مع الكالسيوم والفوسفور للتقييم الكامل"},
    {"if_present":["CBC","Ferritin","Vitamin D3(25 Hydroxy Cholecal.)"],"suggest":["B12","Folic Acid"],"reason":"فحص إجهاد شامل — أضف ب12 وحمض الفوليك"},
]

def get_smart_suggestions(selected_labs_list, price_lookup, phone=None):
    clean_selected = set()
    for entry in selected_labs_list:
        clean_selected.add(entry.split(" — ")[0].strip())
    all_suggestions = {}
    def add_suggestion(name, reason, stype):
        if name not in clean_selected and name not in all_suggestions:
            price = price_lookup.get(name, 0)
            all_suggestions[name] = {"name":name,"reason":reason,"price":price,"type":stype}
    for rule in PANEL_SUGGEST_RULES:
        if not any(t in clean_selected for t in rule["core"]): continue
        for t in [x for x in rule["suggest"] if x not in clean_selected]:
            add_suggestion(t, f"{rule['panel']} — {rule['reason']}", "panel")
    for rule in CLINICAL_RULES:
        if all(t in clean_selected for t in rule["if_present"]):
            for t in rule["suggest"]:
                add_suggestion(t, rule["reason"], "clinical")
    if phone:
        try:
            history = fetch_client_history(phone, limit=5)
            historical_labs = set()
            for v in history:
                for line in (v.get("selected_labs_text","") or "").splitlines():
                    historical_labs.add(line.split(" — ")[0].strip())
            FOLLOWUP_PAIRS = [
                ({"HbA1C"}, ["HbA1C"], "متابعة السكر — آخر مرة كانت في زيارة سابقة"),
                ({"TSH"}, ["TSH","FT3","FT4"], "متابعة الغدة الدرقية — ينصح بالمتابعة الدورية"),
                ({"Vitamin D3(25 Hydroxy Cholecal.)"}, ["Vitamin D3(25 Hydroxy Cholecal.)"], "متابعة فيتامين د — ينصح بالقياس الدوري كل 3 أشهر"),
                ({"Ferritin","Iron (Serum)"}, ["Ferritin","Iron (Serum)","CBC"], "متابعة الحديد — نفس مجموعة التحاليل من قبل"),
            ]
            for hist_trigger, suggest_labs, reason in FOLLOWUP_PAIRS:
                if hist_trigger.issubset(historical_labs):
                    for lab in suggest_labs:
                        if lab not in clean_selected:
                            add_suggestion(lab, f"🕐 {reason}", "history")
        except Exception:
            pass
    bundle_suggestions = []
    for rule in BUNDLE_RULES:
        if any(t in clean_selected for t in rule["trigger"]):
            vd_count = sum(1 for e in selected_labs_list if "Vitamin D3(25 Hydroxy Cholecal.)" in e)
            if vd_count == 1:
                bundle_price = price_lookup.get("Vitamin D3 Couple", 0)
                single_price = price_lookup.get("Vitamin D3(25 Hydroxy Cholecal.)", 400)
                if bundle_price > 0:
                    saving = single_price * 2 - bundle_price
                    bundle_suggestions.append({"name":"Vitamin D3 Couple","note":f"💡 استبدل بـ Couple وادفع {bundle_price} بدل {single_price*2} — توفير {saving} جنيه!","action":"replace","remove":"Vitamin D3(25 Hydroxy Cholecal.)","price":bundle_price,"saving":saving})
                else:
                    bundle_suggestions.append({"name":"Vitamin D3(25 Hydroxy Cholecal.)","note":f"💡 إضافة فيتامين د ثاني للفردين — السعر الإجمالي {single_price + 200} بدل {single_price * 2}","action":"add_discounted","price":200,"saving":single_price - 200})
    return all_suggestions, bundle_suggestions

# ══════════════════════════════════════════════════════════════════════════════
# قائمة الأسعار (كما هي)
# ══════════════════════════════════════════════════════════════════════════════
try:
    from labs_price_list import LABS_DB
    ALL_LABS = [{"name":t["name"],"price":t["price"],"category":cat} for cat, tests in LABS_DB.items() for t in tests]
    LABS_PRICE_LOOKUP = {t["name"]: t["price"] for t in ALL_LABS}
except Exception as e:
    st.error(f"خطأ في استيراد labs_price_list: {e}")
    ALL_LABS = []; LABS_PRICE_LOOKUP = {}

# ══════════════════════════════════════════════════════════════════════════════
# دوال مساعدة
# ══════════════════════════════════════════════════════════════════════════════
def format_date_ar(d):
    if not d: return ""
    if isinstance(d, str):
        try: d = datetime.strptime(d, "%Y-%m-%d").date()
        except: return d
    return f"{d.day} {MONTHS_AR[d.month-1]} {d.year}"

def get_client_tag(phone):
    k = f"_ctag_{phone}"
    if k in st.session_state:
        return st.session_state[k]
    conn = get_connection()
    count = conn.execute(
        "SELECT COUNT(*) FROM visits WHERE phone=? AND status='تمت' AND deleted_at IS NULL", (phone,)
    ).fetchone()[0]
    if count == 0:   tag = "🆕 عميل جديد"
    elif count <= 2: tag = "⭐ عميل منتظم"
    elif count <= 5: tag = "🌟 عميل متكرر"
    else:            tag = "👑 VIP"
    st.session_state[k] = tag
    return tag

def get_client_tag_color(tag):
    return {"🆕 عميل جديد":"#3498DB","⭐ عميل منتظم":"#27AE60",
            "🌟 عميل متكرر":"#F39C12","👑 VIP":"#9B59B6",
            "🏢 Corporate":"#E74C3C"}.get(tag,"#888")

def get_churn_risk(phone):
    conn = get_connection()
    row = conn.execute(
        "SELECT visit_date FROM visits WHERE phone=? AND status='تمت' AND deleted_at IS NULL ORDER BY visit_date DESC LIMIT 1",
        (phone,)
    ).fetchone()
    if not row: return "⚪ لا يوجد تاريخ"
    try:
        last = datetime.strptime(row[0], "%Y-%m-%d")
        days = (datetime.now() - last).days
        if days > 90:   return f"🔴 خطر عالي ({days} يوم)"
        elif days > 45: return f"🟡 متوسط ({days} يوم)"
        else:           return f"🟢 نشط ({days} يوم)"
    except: return "⚪ غير محدد"

def get_today_count(branch_filter=None):
    conn = get_connection()
    today = date.today().isoformat()
    if branch_filter:
        return conn.execute(
            "SELECT COUNT(*) FROM visits WHERE visit_date=? AND branch=? AND status!='ملغية' AND deleted_at IS NULL",
            (today, branch_filter)
        ).fetchone()[0]
    return conn.execute(
        "SELECT COUNT(*) FROM visits WHERE visit_date=? AND status!='ملغية' AND deleted_at IS NULL", (today,)
    ).fetchone()[0]

def get_pending_followups_count():
    today = date.today().isoformat()
    conn = get_connection()
    return conn.execute(
        "SELECT COUNT(*) FROM follow_ups WHERE done=0 AND follow_up_date<=?", (today,)
    ).fetchone()[0]

def get_doctor_workload(doctor_name, visit_date):
    if not doctor_name or doctor_name == "أخرى...": return 0
    conn = get_connection()
    return conn.execute(
        "SELECT COUNT(*) FROM visits WHERE doctor_name=? AND visit_date=? AND status!='ملغية' AND deleted_at IS NULL",
        (doctor_name, str(visit_date))
    ).fetchone()[0]

def cluster_visits_by_area(visits):
    clusters = {}
    for v in visits:
        addr = v.get("address","") or ""
        if "-" in addr:         area = addr.split("-")[0].strip()
        elif "،" in addr:       area = addr.split("،")[0].strip()
        elif "," in addr:       area = addr.split(",")[0].strip()
        else:                   area = addr[:20].strip()
        if not area: area = "غير محدد"
        clusters.setdefault(area, []).append(v)
    return clusters

def make_whatsapp_msg(v, target="internal"):
    lpb = v.get("labs_price_before",0); lpa = v.get("labs_price_after",0)
    tf = v.get("transport_fee",0); total = v.get("total_price",0)
    vdate = format_date_ar(v.get("visit_date","")); vtime = v.get("visit_time","")
    dt_str = vdate + (f" — {vtime}" if vtime else "")
    doc = v.get("doctor_name","غير محدد"); addr = v.get("address","")
    loc = v.get("location_link",""); br = v.get("branch","")
    cname = v.get("name",""); age = v.get("age",""); au = v.get("age_unit","سنة")
    age_s = f"🎂 *العمر:* {age} {au}\n" if age else ""
    lt = v.get("selected_labs_text","")
    if lt.strip():
        labs_lines = "\n".join(f"🧪 {l.strip()}" for l in lt.splitlines() if l.strip()) + "\n"
    else:
        labs_lines = "🚫 لا توجد تحاليل\n"
    loc_line = f"📍 *الموقع:* {loc}\n" if loc else ""
    br_line  = f"🏥 *الفرع:* {br}\n"   if br  else ""
    status   = v.get("status","مجدولة")
    if target == "client":
        return (f"🟠 *Orange Lab Home Visit Management System*\n🏠 أهلاً بك {cname}\n━━━━━━━━━━━━━━\n"
                f"👨‍⚕️ *الدكتور القائم بالزيارة:* {doc}\n📅 *موعد الزيارة:* {dt_str}\n━━━━━━━━━━━━━━\n"
                f"📍 *عنوان الزيارة:*\n{addr}\n{loc_line}{br_line}━━━━━━━━━━━━━━\n"
                f"🧪 *التحاليل المطلوبة:*\n{labs_lines}━━━━━━━━━━━━━━\n"
                f"💰 *السعر قبل الخصم:* {lpb} جنيه\n💰 *السعر بعد الخصم:* {lpa} جنيه\n"
                f"🚗 *بدل الانتقال:* {tf} جنيه\n💵 *الإجمالي المطلوب:* {total} جنيه\n━━━━━━━━━━━━━━\n"
                f"✏️ *برجاء تأكيد حجزك بالرد برقم:*\n  1 - تأكيد الزيارة\n  2 - تأجيل الزيارة\n  3 - إلغاء الزيارة\n\n"
                f"شكراً لثقتكم 🧡 *معمل أورانج لاب*")
    elif target == "group":
        return (f"🟠 *زيارة منزلية*\n━━━━━━━━━━━━━━\n"
                f"👨‍⚕️ *الدكتور القائم بالزيارة:* {doc}\n📅 *الموعد:* {dt_str}")
    else:
        notes_line = f"📝 *ملاحظات:* {v.get('notes','')}\n" if v.get("notes") else ""
        return (f"🟠 *Orange Lab Home Visit Management System*\n━━━━━━━━━━━━━━\n"
                f"👤 *الاسم:* {v['name']}\n{age_s}📞 *التليفون:* {v.get('phone','')}\n"
                f"📅 *الموعد:* {dt_str}\n👨‍⚕️ *دكتور الزيارة:* {doc}\n"
                f"🏥 *الفرع:* {br}\n🔖 *الحالة:* {STATUS_ICONS.get(status,'')} {status}\n━━━━━━━━━━━━━━\n"
                f"📍 *العنوان:* {addr}\n{loc_line}━━━━━━━━━━━━━━\n"
                f"🧪 *التحاليل المطلوبة:*\n{labs_lines}━━━━━━━━━━━━━━\n"
                f"💰 *السعر قبل الخصم:* {lpb} جنيه\n💰 *السعر بعد الخصم:* {lpa} جنيه\n"
                f"🚗 *بدل الانتقال:* {tf} جنيه\n💵 *الإجمالي:* {total} جنيه\n━━━━━━━━━━━━━━\n{notes_line}")

def whatsapp_link(msg, phone=None):
    encoded = urllib.parse.quote(msg, encoding="utf-8")
    if phone:
        p = phone.strip().replace(" ","").replace("-","").replace("+","")
        if p.startswith("0"):   p = "20" + p[1:]
        elif not p.startswith("20"): p = "20" + p
        return f"https://wa.me/{p}?text={encoded}"
    return f"https://wa.me/?text={encoded}"

def generate_visit_print_html(v):
    lt = v.get("selected_labs_text","")
    labs_rows = "".join(f"<tr><td style='padding:6px 10px;border-bottom:1px solid #eee;'>🔹 {l.strip()}</td></tr>"
        for l in lt.splitlines() if l.strip()) if lt.strip() else "<tr><td>لا توجد تحاليل</td></tr>"
    status = v.get("status","مجدولة")
    s_color = STATUS_COLORS.get(status,"#888"); s_icon = STATUS_ICONS.get(status,"")
    loc_html = f'<p style="font-size:12px;color:#FF6B00;">🗺️ {v.get("location_link","")}</p>' if v.get("location_link") else ""
    notes_html = f'<div class="section"><div class="section-title">📌 ملاحظات</div><p style="font-size:13px;">{v.get("notes","")}</p></div>' if v.get("notes") else ""
    pay_status = v.get("payment_status","غير مدفوع")
    pay_color = PAYMENT_COLORS.get(pay_status,"#888")
    return f"""<!DOCTYPE html><html dir="rtl">
<head><meta charset="UTF-8">
<link href="https://fonts.googleapis.com/css2?family=Cairo:wght@400;700;800&display=swap" rel="stylesheet">
<style>
body{{font-family:'Cairo',sans-serif;margin:30px;color:#222;background:#fff;}}
.header{{background:linear-gradient(90deg,#FF6B00,#FF9A3C);color:#fff;border-radius:12px;padding:16px 22px;display:flex;justify-content:space-between;align-items:center;margin-bottom:20px;}}
.header h2{{margin:0;font-size:20px;}} .header span{{font-size:13px;opacity:.85;}}
.section{{margin-bottom:18px;}} .section-title{{color:#FF6B00;font-weight:800;font-size:14px;border-right:4px solid #FF6B00;padding-right:10px;margin-bottom:10px;}}
.row{{display:flex;justify-content:space-between;padding:7px 0;border-bottom:1px solid #f0f0f0;font-size:13px;}}
.label{{color:#888;}} .value{{font-weight:700;}}
.status-badge{{display:inline-block;background:{s_color};color:#fff;border-radius:20px;padding:3px 14px;font-size:12px;font-weight:700;}}
.pay-badge{{display:inline-block;background:{pay_color};color:#fff;border-radius:20px;padding:3px 14px;font-size:12px;font-weight:700;}}
.price-box{{background:linear-gradient(135deg,#FF6B00,#FF9A3C);border-radius:12px;padding:14px 18px;color:#fff;margin-top:16px;}}
.price-row{{display:flex;justify-content:space-between;font-size:13px;margin-bottom:6px;}}
.price-total{{display:flex;justify-content:space-between;font-size:17px;font-weight:800;border-top:2px solid rgba(255,255,255,.3);padding-top:8px;margin-top:4px;}}
table{{width:100%;border-collapse:collapse;font-size:13px;}}
.footer{{text-align:center;margin-top:30px;color:#aaa;font-size:11px;border-top:1px solid #eee;padding-top:12px;}}
</style></head><body>
<div class="header"><h2>🟠 Orange Lab Home Visit Management System</h2><span>📅 {format_date_ar(v.get('visit_date',''))}</span></div>
<div class="section"><div class="section-title">👤 بيانات العميل</div>
<div class="row"><span class="label">الاسم</span><span class="value">{v['name']}</span></div>
<div class="row"><span class="label">السن</span><span class="value">{v.get('age','')} {v.get('age_unit','سنة')}</span></div>
<div class="row"><span class="label">التليفون</span><span class="value">{v.get('phone','')}</span></div>
<div class="row"><span class="label">الموعد</span><span class="value">{format_date_ar(v.get('visit_date',''))} — {v.get('visit_time','')}</span></div>
<div class="row"><span class="label">الدكتور</span><span class="value">{v.get('doctor_name','')}</span></div>
<div class="row"><span class="label">الفرع</span><span class="value">{v.get('branch','')}</span></div>
<div class="row"><span class="label">الحالة</span><span class="value"><span class="status-badge">{s_icon} {status}</span></span></div>
<div class="row"><span class="label">الدفع</span><span class="value"><span class="pay-badge">{PAYMENT_ICONS.get(pay_status,'')} {pay_status}</span></span></div>
</div>
<div class="section"><div class="section-title">📍 العنوان</div>
<p style="margin:6px 0;font-size:13px;">{v.get('address','')}</p>{loc_html}</div>
<div class="section"><div class="section-title">🧪 التحاليل المطلوبة</div>
<table><tbody>{labs_rows}</tbody></table></div>
{notes_html}
<div class="price-box">
<div class="price-row"><span>⭐ السعر قبل الخصم</span><span>{v.get('labs_price_before',0)} جنيه</span></div>
<div class="price-row"><span>⭐ السعر بعد الخصم</span><span>{v.get('labs_price_after',0)} جنيه</span></div>
<div class="price-row"><span>🚗 بدل الانتقال</span><span>{v.get('transport_fee',0)} جنيه</span></div>
<div class="price-total"><span>💵 الإجمالي</span><span>{v.get('total_price',0)} جنيه</span></div>
</div>
<div class="footer">Orange Lab Home Visit Management System — Developed by Dr / Hussein Ali 2026</div>
</body></html>"""

# ══════════════════════════════════════════════════════════════════════════════
# CSS
# ══════════════════════════════════════════════════════════════════════════════
def inject_css():
    css = """
    <link href="https://fonts.googleapis.com/css2?family=Cairo:wght@400;600;700;800&display=swap" rel="stylesheet">
    <style>
      html, body, [class*="css"] { font-family: 'Cairo', sans-serif !important; direction: rtl; }
      .main { background: #fff8f0; }
      .block-container { padding-top: 0.5rem !important; max-width: 680px; }
      .ohv-header { background: linear-gradient(135deg,#FF6B00,#FF9A3C); border-radius:16px; padding:18px 24px;
        display:flex; align-items:center; justify-content:space-between; margin-bottom:18px;
        box-shadow: 0 6px 24px rgba(255,107,0,0.35); }
      .ohv-header-left { display:flex; flex-direction:column; gap:3px; }
      .ohv-header h1 { color:#fff; margin:0; font-size:26px; font-weight:800; letter-spacing:-0.5px; line-height:1.1; }
      .ohv-header-dev { color:rgba(255,255,255,0.82); font-size:12px; font-weight:500; margin-top:2px; }
      .ohv-header-date { color:rgba(255,255,255,0.85); font-size:12px; text-align:right; }
      .stat-grid { display:flex; gap:10px; margin-bottom:18px; flex-wrap:wrap; }
      .stat-box { flex:1; min-width:80px; background:#fff; border-radius:14px; padding:12px;
        text-align:center; border:1px solid #ffe8d1; box-shadow:0 2px 10px rgba(0,0,0,0.05); }
      .stat-num { font-size:24px; font-weight:800; color:#FF6B00; }
      .stat-label { font-size:10px; color:#aaa; margin-top:2px; }
      .visit-card { background:#fff; border-radius:14px; padding:14px; margin-bottom:10px;
        border:1px solid #ffe8d1; box-shadow:0 2px 10px rgba(0,0,0,0.05); }
      .visit-name { font-size:15px; font-weight:700; color:#222; margin-top:6px; }
      .visit-meta { font-size:12px; color:#888; margin-top:4px; }
      .visit-badge { background:#fff3e6; color:#FF6B00; border-radius:8px;
        padding:3px 10px; font-size:12px; font-weight:700; float:left; }
      .status-badge { display:inline-block; border-radius:20px; padding:3px 12px;
        font-size:11px; font-weight:700; color:#fff; margin-right:4px; }
      .pay-badge { display:inline-block; border-radius:20px; padding:2px 10px;
        font-size:10px; font-weight:700; color:#fff; margin-right:2px; }
      .price-box { background:linear-gradient(135deg,#FF6B00,#FF9A3C); border-radius:16px;
        padding:16px 20px; color:#fff; margin-bottom:14px; }
      .price-row { display:flex; justify-content:space-between; font-size:14px; margin-bottom:7px; }
      .price-total { display:flex; justify-content:space-between; font-size:19px; font-weight:800;
        border-top:2px solid rgba(255,255,255,0.3); padding-top:9px; margin-top:5px; }
      .wa-btn { display:block; padding:11px 16px; border-radius:12px; color:#fff !important;
        font-weight:700; font-size:13px; text-decoration:none; text-align:center;
        font-family:'Cairo',sans-serif; margin-bottom:8px; }
      .wa-client { background:#25D366; } .wa-share { background:#128C7E; } .wa-group { background:#075E54; }
      .detail-row { display:flex; justify-content:space-between; padding:8px 0;
        border-bottom:1px solid #f5f5f5; font-size:13px; }
      .detail-label { color:#888; }
      .detail-value { font-weight:600; color:#222; max-width:58%; text-align:left; }
      .repeat-banner { background:#fff8f0; border:2px dashed #FF9A3C; border-radius:14px;
        padding:12px; text-align:center; margin-top:12px; color:#FF6B00; font-weight:700; font-size:14px; }
      .section-title { font-size:14px; font-weight:700; color:#FF6B00;
        border-right:4px solid #FF6B00; padding-right:10px; margin-bottom:10px; }
      .history-card { background:#f9f9f9; border-radius:10px; padding:10px 14px;
        margin-bottom:8px; border-right:4px solid #FF9A3C; font-size:13px; }
      .today-header { background:linear-gradient(90deg,#27AE60,#2ECC71); border-radius:14px;
        padding:12px 18px; color:#fff; font-weight:800; font-size:15px;
        margin-bottom:14px; text-align:center; }
      .area-header { background:linear-gradient(90deg,#2980B9,#3498DB); border-radius:10px;
        padding:8px 14px; color:#fff; font-weight:700; font-size:13px; margin:10px 0 6px 0; }
      .fu-card { background:#fff; border-radius:12px; padding:12px 16px; margin-bottom:8px;
        border:1px solid #ffe8d1; border-right:4px solid #E67E22; }
      .fu-done { border-right:4px solid #27AE60; opacity:0.7; }
      .fu-overdue { border-right:4px solid #E74C3C; background:#FFF5F5; }
      .notif-banner { background:linear-gradient(90deg,#E67E22,#F39C12); border-radius:12px;
        padding:10px 16px; color:#fff; font-weight:700; font-size:13px; margin-bottom:12px;
        display:flex; align-items:center; gap:10px; }
      .client-profile-header { background:linear-gradient(135deg,#1a1a2e,#16213e);
        border-radius:16px; padding:20px; color:#fff; margin-bottom:18px; }
      div[data-testid="stButton"] button { font-family:'Cairo',sans-serif !important; font-weight:700 !important; border-radius:12px !important; }
      div[data-testid="stTextInput"] label, div[data-testid="stNumberInput"] label,
      div[data-testid="stDateInput"] label, div[data-testid="stTextArea"] label,
      div[data-testid="stMultiSelect"] label, div[data-testid="stSelectbox"] label {
        font-family:'Cairo',sans-serif !important; font-weight:600 !important; color:#555 !important; }
      div[data-testid="stTextInput"] input, div[data-testid="stNumberInput"] input,
      div[data-testid="stTextArea"] textarea, div[data-testid="stDateInput"] input {
        background-color:#FFF3E8 !important; border:1.5px solid #FFBB80 !important;
        border-radius:8px !important; color:#222 !important; }
      div[data-testid="stTextInput"] input:focus, div[data-testid="stNumberInput"] input:focus,
      div[data-testid="stTextArea"] textarea:focus, div[data-testid="stDateInput"] input:focus {
        background-color:#FFE8CC !important; border:2px solid #FF6B00 !important;
        box-shadow:0 0 0 3px rgba(255,107,0,0.15) !important; outline:none !important; }
      div[data-testid="stSelectbox"] > div > div {
        background-color:#FFF3E8 !important; border:1.5px solid #FFBB80 !important; border-radius:8px !important; }
      #MainMenu { visibility: hidden; } footer { visibility: hidden; } header { visibility: hidden; }
      @media print {
        body * { visibility: hidden; }
        #printable-report, #printable-report * { visibility: visible; }
        #printable-report { position:absolute; left:0; top:0; width:100%; }
        .no-print { display: none; }
      }
    </style>"""
    st.markdown(css, unsafe_allow_html=True)

inject_css()

# ══════════════════════════════════════════════════════════════════════════════
# visit card html
# ══════════════════════════════════════════════════════════════════════════════
def visit_card_html(v):
    total     = v.get("total_price", 0)
    vdate     = format_date_ar(v.get("visit_date",""))
    vtime     = v.get("visit_time","")
    addr      = (v.get("address","") or "")
    addr_short = addr[:38] + ("..." if len(addr)>38 else "")
    lc        = len(v.get("selected_labs_text","").splitlines()) if v.get("selected_labs_text") else 0
    doc_show  = f" | 👨‍⚕️ {v.get('doctor_name','')}" if v.get("doctor_name") else ""
    br_show   = f" | 🏥 {v.get('branch','')}"        if v.get("branch")      else ""
    age       = v.get("age",""); au = v.get("age_unit","سنة")
    age_disp  = f"🎂 {age} {au}" if age else ""
    status    = v.get("status","مجدولة")
    sc        = STATUS_COLORS.get(status,"#888"); si = STATUS_ICONS.get(status,"")
    tag_auto  = get_client_tag(v.get("phone",""))
    tag_color = get_client_tag_color(tag_auto)
    rating    = int(v.get("rating", 0) or 0)
    stars     = "⭐" * rating if rating else ""
    pay_status = v.get("payment_status","غير مدفوع")
    pay_color  = PAYMENT_COLORS.get(pay_status,"#888")
    pay_icon   = PAYMENT_ICONS.get(pay_status,"🔴")
    return (
        f'<div class="visit-card">'
        f'<span class="visit-badge">{total:,} جنيه</span>'
        f'<span class="status-badge" style="background:{sc}">{si} {status}</span>'
        f'<span class="pay-badge" style="background:{pay_color}">{pay_icon} {pay_status}</span>'
        f'<span style="background:{tag_color};color:#fff;border-radius:8px;'
        f'padding:2px 8px;font-size:10px;font-weight:700;margin-right:4px;">{tag_auto}</span>'
        f'<div class="visit-name">👤 {v["name"]} {stars}</div>'
        f'<div class="visit-meta">📞 {v.get("phone","")} &nbsp;|&nbsp; 📅 {vdate} {vtime} &nbsp; {age_disp}</div>'
        f'<div class="visit-meta">📍 {addr_short}</div>'
        f'<div class="visit-meta" style="margin-top:5px">🧪 {lc} تحليل{doc_show}{br_show}</div>'
        f'</div>'
    )

# ══════════════════════════════════════════════════════════════════════════════
# Navigation
# ══════════════════════════════════════════════════════════════════════════════
def go(page, prefill=None, visit_id=None, client_phone=None):
    if page == "new" and (prefill is None or not prefill.get("_edit")):
        st.session_state.pop("added_labs_new_visit", None)
    st.session_state.page = page
    if prefill      is not None: st.session_state.prefill             = prefill
    if visit_id     is not None: st.session_state.selected_id         = visit_id
    if client_phone is not None: st.session_state.selected_client_phone = client_phone
    st.rerun()

# ── Header ──
st.markdown(f"""<div class="ohv-header">
  <div class="ohv-header-left">
    <h1>🟠 Orange Lab HVMS</h1>
    <div class="ohv-header-dev">Developed by: Dr / Hussein Ali &nbsp;for&nbsp; Orange Lab 🍊</div>
  </div>
  <div class="ohv-header-date">📅 {format_date_ar(date.today())}</div>
</div>""", unsafe_allow_html=True)

# ── Notification Banner: follow-ups pending ──
_pf_count = get_pending_followups_count()
if _pf_count > 0:
    st.markdown(f'<div class="notif-banner">⏰ عندك <b>{_pf_count}</b> متابعة معلقة اليوم أو متأخرة — '
                f'<a href="#" style="color:#fff;text-decoration:underline;">اضغط على المتابعات</a></div>',
                unsafe_allow_html=True)

# ── Today count for button label ──
_utype = st.session_state.user_type
if _utype == "diamond":   _today_branch = "Diamond"
elif _utype == "lacite":  _today_branch = "La Cite"
else:                     _today_branch = None
_today_n = get_today_count(_today_branch)
_today_label = f"📅 اليوم ({_today_n})" if _today_n > 0 else "📅 اليوم"

# ── Nav Buttons ──
if _utype == "admin":
    nc1,nc2,nc3,nc4,nc5,nc6,nc7,nc8 = st.columns([2,2,2,2,2,2,2,1])
    with nc5:
        if st.button("📊 Dashboard", use_container_width=True): go("dashboard")
    with nc6:
        if st.button("👨‍⚕️ أطباء", use_container_width=True): go("manage_doctors")
    with nc7:
        fu_label = f"⏰ متابعات ({_pf_count})" if _pf_count>0 else "⏰ متابعات"
        if st.button(fu_label, use_container_width=True): go("follow_ups")
    with nc8:
        if st.button("🚪", help="تسجيل الخروج", use_container_width=True):
            st.session_state.authenticated = False; st.rerun()
    with nc1:
        if st.button("🏠 الرئيسية", use_container_width=True): go("home")
    with nc2:
        if st.button("➕ جديدة", use_container_width=True): go("new", prefill={})
    with nc3:
        if st.button(_today_label, use_container_width=True): go("today")
    with nc4:
        if st.button("📈 تقارير", use_container_width=True): go("reports")
elif _utype in ["diamond","lacite"]:
    nc1,nc2,nc3,nc4,nc5,nc6 = st.columns([2,2,2,2,2,1])
    with nc5:
        fu_label = f"⏰ متابعات ({_pf_count})" if _pf_count>0 else "⏰ متابعات"
        if st.button(fu_label, use_container_width=True): go("follow_ups")
    with nc6:
        if st.button("🚪", help="تسجيل الخروج", use_container_width=True):
            st.session_state.authenticated = False; st.rerun()
    with nc1:
        if st.button("🏠 الرئيسية", use_container_width=True): go("home")
    with nc2:
        if st.button("➕ جديدة", use_container_width=True): go("new", prefill={})
    with nc3:
        if st.button(_today_label, use_container_width=True): go("today")
    with nc4:
        if st.button("📈 تقارير", use_container_width=True): go("reports")
else:
    nc1, nc2 = st.columns([4,1])
    with nc1:
        if st.button("➕ زيارة جديدة", use_container_width=True): go("new", prefill={})
    with nc2:
        if st.button("🚪", help="تسجيل الخروج", use_container_width=True):
            st.session_state.authenticated = False; st.rerun()

st.markdown("---")
# ══════════════════════════════════════════════════════════════════════════════
# صفحة الرئيسية
# ══════════════════════════════════════════════════════════════════════════════
if st.session_state.page == "home":
    if st.session_state.user_type not in ["admin","diamond","lacite"]:
        st.info("ليس لديك صلاحية عرض بيانات الزيارات."); st.stop()
    conn = get_connection()
    all_doctors  = [r[0] for r in conn.execute("SELECT DISTINCT doctor_name FROM visits WHERE doctor_name!='' AND deleted_at IS NULL").fetchall()]
    all_branches = [r[0] for r in conn.execute("SELECT DISTINCT branch FROM visits WHERE deleted_at IS NULL").fetchall()]
    for lst in [all_branches, all_doctors]:
        if "الكل" not in lst: lst.insert(0, "الكل")
    st.markdown("### تصفية الزيارات")
    cf1,cf2,cf3,cf4,cf5 = st.columns(5)
    with cf1:
        if st.session_state.user_type == "diamond":
            selected_branch = "Diamond"; st.selectbox("الفرع",["Diamond"],disabled=True)
        elif st.session_state.user_type == "lacite":
            selected_branch = "La Cite"; st.selectbox("الفرع",["La Cite"],disabled=True)
        else:
            selected_branch = st.selectbox("الفرع", all_branches, index=0)
    with cf2:
        selected_doctor = st.selectbox("الدكتور", all_doctors, index=0)
    with cf3:
        status_opts     = ["الكل"] + STATUS_OPTIONS
        selected_status = st.selectbox("الحالة", status_opts, index=0)
    with cf4:
        pay_opts = ["الكل"] + PAYMENT_STATUS_OPTIONS
        selected_pay = st.selectbox("الدفع", pay_opts, index=0)
    with cf5:
        search_query = st.text_input("🔍 بحث", value=st.session_state.search_q, placeholder="اسم، تليفون، عنوان، طبيب، رقم الزيارة")
        st.session_state.search_q = search_query
    filters = {}
    if selected_branch != "الكل": filters["branch"] = selected_branch
    if selected_doctor != "الكل": filters["doctor"] = selected_doctor
    if selected_status != "الكل": filters["status"] = selected_status
    if selected_pay    != "الكل": filters["payment_status"] = selected_pay
    if search_query:               filters["search"] = search_query

    # Pagination
    page = st.session_state.current_page
    page_size = st.session_state.page_size
    visits, total_visits = fetch_visits(filters, page=page, page_size=page_size)
    st.session_state.total_visits = total_visits

    today_s = date.today().isoformat()
    bf_kpi  = "Diamond" if st.session_state.user_type=="diamond" else "La Cite" if st.session_state.user_type=="lacite" else None
    all_vs  = fetch_visits({"branch":bf_kpi} if bf_kpi else {}, page=None, page_size=None)[0]  # جلب الكل للإحصائيات
    t_today = sum(1 for v in all_vs if v.get("visit_date")==today_s)
    t_rev   = sum(v.get("total_price",0) for v in all_vs if v.get("status")!="ملغية")
    t_done  = sum(1 for v in all_vs if v.get("status")=="تمت")
    t_unpaid = sum(1 for v in all_vs if v.get("payment_status","غير مدفوع")=="غير مدفوع" and v.get("status")!="ملغية")
    st.markdown(f"""<div class="stat-grid">
      <div class="stat-box"><div class="stat-num">{len(all_vs)}</div><div class="stat-label">إجمالي الزيارات</div></div>
      <div class="stat-box"><div class="stat-num">{t_today}</div><div class="stat-label">زيارات اليوم</div></div>
      <div class="stat-box"><div class="stat-num" style="color:#27AE60">{t_done}</div><div class="stat-label">تمت ✅</div></div>
      <div class="stat-box"><div class="stat-num" style="font-size:16px">{t_rev:,.0f}</div><div class="stat-label">الإيراد (جنيه)</div></div>
      <div class="stat-box"><div class="stat-num" style="color:#E74C3C">{t_unpaid}</div><div class="stat-label">غير مدفوع 🔴</div></div>
    </div>""", unsafe_allow_html=True)
    if st.session_state.user_type in ["admin","diamond","lacite"]:
        with st.expander("📤 تصدير إلى Excel", expanded=False):
            st.markdown('<div style="font-size:13px;font-weight:700;color:#FF6B00;margin-bottom:8px">📅 اختر فترة التصدير</div>', unsafe_allow_html=True)
            exp_c1, exp_c2 = st.columns(2)
            with exp_c1:
                exp_from = st.date_input("من تاريخ", value=None, key="exp_from", help="اكتب أو اختر تاريخ البداية")
            with exp_c2:
                exp_to   = st.date_input("إلى تاريخ", value=date.today(), key="exp_to")
            if st.session_state.user_type == "admin":
                exp_branch_sel = st.selectbox("الفرع", ["كل الفروع","La Cite","Diamond"], key="exp_branch_sel")
                bf_exp = None if exp_branch_sel == "كل الفروع" else exp_branch_sel
                btn_label = f"📤 تصدير ({exp_branch_sel})"
            elif st.session_state.user_type == "lacite":
                bf_exp = "La Cite"; btn_label = "📤 تصدير زيارات La Cite"
            else:
                bf_exp = "Diamond"; btn_label = "📤 تصدير زيارات Diamond"
            if st.button(btn_label, use_container_width=True, key="btn_export_home"):
                if not exp_from or not exp_to:
                    st.error("⚠️ من فضلك اختر تاريخ البداية والنهاية")
                elif exp_from > exp_to:
                    st.error("⚠️ تاريخ البداية أكبر من تاريخ النهاية!")
                else:
                    df_ex, path_ex = export_to_excel(branch_filter=bf_exp, date_from=exp_from.isoformat(), date_to=exp_to.isoformat())
                    if df_ex.empty:
                        st.warning("لا توجد زيارات في هذه الفترة.")
                    else:
                        period_label = f"{exp_from} → {exp_to}"
                        prefix = "diamond" if bf_exp=="Diamond" else "lacite" if bf_exp=="La Cite" else "visits"
                        fname_dl = f"{prefix}_{exp_from}_{exp_to}.xlsx"
                        with open(path_ex,"rb") as fh:
                            st.download_button(f"📥 تحميل ({period_label}) — {len(df_ex)} زيارة", data=fh, file_name=fname_dl,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_home_final")
    if st.session_state.user_type == "admin":
        uf = st.file_uploader("📥 استيراد من Excel", type=["xlsx"], key="import_excel")
        if uf:
            count_imported, count_updated = import_from_excel(uf)
            if count_updated > 0:
                st.success(f"✅ تم استيراد {count_imported} زيارة جديدة، وتحديث {count_updated} زيارة موجودة!")
            else:
                st.success(f"✅ تم استيراد {count_imported} زيارة جديدة!")
            st.rerun()
    st.markdown("---")
    if not visits:
        st.info("لا توجد زيارات تطابق التصفية.")
    else:
        for v in visits:
            st.markdown(visit_card_html(v), unsafe_allow_html=True)
            vc1,vc2 = st.columns([5,2])
            with vc1:
                if st.button(f"📂 فتح {v['name']}", key=f"o_{v['id']}", use_container_width=True):
                    go("detail", visit_id=v["id"])
            with vc2:
                if st.button(f"👤 بروفايل", key=f"cp_{v['id']}", use_container_width=True):
                    go("client_profile", client_phone=v.get("phone",""))
        # Pagination controls
        total_pages = (total_visits + page_size - 1) // page_size
        if total_pages > 1:
            col1, col2, col3, col4, col5 = st.columns([1,1,2,1,1])
            with col1:
                if st.button("⬅️ السابق", disabled=(page<=1), use_container_width=True):
                    st.session_state.current_page = max(1, page-1); st.rerun()
            with col2:
                st.write(f"الصفحة {page} من {total_pages}")
            with col3:
                page_input = st.number_input("اذهب إلى صفحة", min_value=1, max_value=total_pages, value=page, key="page_input", label_visibility="collapsed")
                if page_input != page:
                    st.session_state.current_page = page_input; st.rerun()
            with col4:
                if st.button("التالي ➡️", disabled=(page>=total_pages), use_container_width=True):
                    st.session_state.current_page = min(total_pages, page+1); st.rerun()
            with col5:
                st.write(f"إجمالي {total_visits} زيارة")
    st.markdown("""
    <div style="text-align:center;margin-top:50px;padding-top:20px;border-top:2px solid #FF6B00;
                color:#333;font-size:14px;font-weight:600;">
      Orange Lab Home Visit Management System<br>Developed by <b>Dr / Hussein Ali</b> 2026
      For <span style="color:#FF6B00;">Orange Lab 🍊</span>
    </div>""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# صفحة زيارات اليوم
# ══════════════════════════════════════════════════════════════════════════════
elif st.session_state.page == "today":
    if st.session_state.user_type not in ["admin","diamond","lacite"]:
        st.error("غير مصرح."); st.stop()
    f = {"date_exact": date.today().isoformat()}
    if st.session_state.user_type == "diamond":   f["branch"] = "Diamond"
    elif st.session_state.user_type == "lacite":  f["branch"] = "La Cite"
    today_visits, _ = fetch_visits(f, page=None, page_size=None)  # جلب الكل لليوم
    today_visits.sort(key=lambda v: v.get("visit_time","") or "")
    st.markdown(f'<div class="today-header">📅 زيارات اليوم — {format_date_ar(date.today())} ({len(today_visits)} زيارة)</div>', unsafe_allow_html=True)
    done_t    = sum(1 for v in today_visits if v.get("status")=="تمت")
    pending_t = sum(1 for v in today_visits if v.get("status") in ["مجدولة","في الطريق"])
    rev_t     = sum(v.get("total_price",0) for v in today_visits if v.get("status")!="ملغية")
    unpaid_t  = sum(1 for v in today_visits if v.get("payment_status","غير مدفوع")=="غير مدفوع" and v.get("status")!="ملغية")
    st.markdown(f"""<div class="stat-grid">
      <div class="stat-box"><div class="stat-num">{len(today_visits)}</div><div class="stat-label">إجمالي اليوم</div></div>
      <div class="stat-box"><div class="stat-num" style="color:#27AE60">{done_t}</div><div class="stat-label">تمت ✅</div></div>
      <div class="stat-box"><div class="stat-num" style="color:#F39C12">{pending_t}</div><div class="stat-label">متبقية 🕐</div></div>
      <div class="stat-box"><div class="stat-num" style="font-size:15px">{rev_t:,.0f}</div><div class="stat-label">إيراد اليوم</div></div>
      <div class="stat-box"><div class="stat-num" style="color:#E74C3C">{unpaid_t}</div><div class="stat-label">غير مدفوع 🔴</div></div>
    </div>""", unsafe_allow_html=True)
    if not today_visits:
        st.info("لا توجد زيارات مجدولة اليوم.")
    else:
        tc1, tc2 = st.columns(2)
        with tc1:
            view_mode = st.radio("عرض", ["📋 قائمة","🗂️ حسب المنطقة"], horizontal=True, key="today_view")
        with tc2:
            tomorrow = (date.today() + timedelta(days=1)).isoformat()
            tomorrow_visits, _ = fetch_visits({**({k:v for k,v in f.items() if k!="date_exact"}), "date_exact": tomorrow}, page=None, page_size=None)
            if tomorrow_visits:
                bulk_msg = f"🟠 *تذكير زيارات غد — {format_date_ar(date.today()+timedelta(1))}*\n━━━━━━━━━━━━━━\n"
                for i, tv in enumerate(tomorrow_visits, 1):
                    bulk_msg += f"{i}. 👤 {tv['name']} — 📅 {tv.get('visit_time','')} — 📍 {(tv.get('address','') or '')[:30]}\n"
                bulk_msg += f"\nإجمالي: {len(tomorrow_visits)} زيارة — Orange Lab 🟠"
                st.markdown(f'<a href="{whatsapp_link(bulk_msg)}" target="_blank" class="wa-btn" style="background:#E67E22;padding:7px 14px;font-size:12px;">📤 إرسال تذكير غد ({len(tomorrow_visits)} زيارة)</a>', unsafe_allow_html=True)
        if view_mode == "📋 قائمة":
            for v in today_visits:
                st.markdown(visit_card_html(v), unsafe_allow_html=True)
                tc1b,tc2b,tc3b = st.columns([3,2,2])
                with tc1b:
                    if st.button(f"📂 فتح {v['name']}", key=f"td_{v['id']}", use_container_width=True):
                        go("detail", visit_id=v["id"])
                with tc2b:
                    cur_idx = STATUS_OPTIONS.index(v.get("status","مجدولة")) if v.get("status") in STATUS_OPTIONS else 0
                    new_stat = st.selectbox("", STATUS_OPTIONS, index=cur_idx, key=f"st_{v['id']}", label_visibility="collapsed")
                    if new_stat != v.get("status","مجدولة"):
                        update_status_only(v["id"], new_stat); st.rerun()
                with tc3b:
                    cur_pay = v.get("payment_status","غير مدفوع")
                    pi = PAYMENT_STATUS_OPTIONS.index(cur_pay) if cur_pay in PAYMENT_STATUS_OPTIONS else 0
                    new_pay = st.selectbox("", PAYMENT_STATUS_OPTIONS, index=pi, key=f"pay_{v['id']}", label_visibility="collapsed")
                    if new_pay != cur_pay:
                        update_payment(v["id"], new_pay, v.get("payment_method",""), v.get("paid_amount",0), date.today().isoformat()); st.rerun()
        else:
            clusters = cluster_visits_by_area(today_visits)
            for area, area_visits in clusters.items():
                st.markdown(f'<div class="area-header">📍 {area} — {len(area_visits)} زيارة</div>', unsafe_allow_html=True)
                for v in area_visits:
                    st.markdown(visit_card_html(v), unsafe_allow_html=True)
                    if st.button(f"📂 فتح {v['name']}", key=f"tdg_{v['id']}", use_container_width=True):
                        go("detail", visit_id=v["id"])

# ══════════════════════════════════════════════════════════════════════════════
# صفحة زيارة جديدة / تعديل
# ══════════════════════════════════════════════════════════════════════════════
elif st.session_state.page == "new":
    pf      = st.session_state.prefill or {}
    is_edit = pf.get("_edit", False)
    st.markdown(f"### {'✏️ تعديل الزيارة' if is_edit else '➕ زيارة جديدة'}")
    st.markdown('<div class="section-title">👤 البيانات الشخصية</div>', unsafe_allow_html=True)
    name  = st.text_input("الاسم الكامل *", value=pf.get("name",""))
    nc1,nc2 = st.columns(2)
    with nc1:
        age = st.number_input("العمر *", 0, 120, int(pf.get("age",0) or 0))
    with nc2:
        au_opts = ["سنة","شهر"]; cur_au = pf.get("age_unit","سنة")
        if cur_au not in au_opts: cur_au = "سنة"
        age_unit = st.radio("الوحدة", au_opts, index=au_opts.index(cur_au), horizontal=True)
    phone = st.text_input("رقم التليفون *", value=pf.get("phone",""), placeholder="01xxxxxxxxx")
    if phone and len(phone) >= 10 and not is_edit:
        prev_visits = fetch_client_history(phone, limit=3)
        if prev_visits:
            tag_auto  = get_client_tag(phone); tag_color = get_client_tag_color(tag_auto)
            last_v    = prev_visits[0]
            last_date = format_date_ar(last_v.get("visit_date",""))
            last_stat = last_v.get("status",""); n_total = len(prev_visits)
            churn     = get_churn_risk(phone)
            st.markdown(f"""<div style="background:#FFF3CD;border:2px solid #F39C12;border-radius:12px;padding:12px 16px;margin:8px 0;direction:rtl;">
              <div style="font-weight:800;color:#E67E22;font-size:14px;margin-bottom:6px;">⚠️ هذا العميل موجود في النظام</div>
              <div style="font-size:13px;color:#333;line-height:1.8;">
                👤 <b>{last_v.get('name','')}</b> &nbsp;|&nbsp;
                <span style="background:{tag_color};color:#fff;border-radius:8px;padding:2px 8px;font-size:11px;">{tag_auto}</span><br>
                📋 إجمالي الزيارات: <b>{n_total}</b><br>
                📅 آخر زيارة: <b>{last_date}</b> — {last_stat}<br>
                🩺 آخر طبيب: {last_v.get('doctor_name','')}<br>
                📊 نشاط العميل: {churn}<br>
                🧪 آخر تحاليل: {last_v.get('selected_labs_text','')[:50] if last_v.get('selected_labs_text') else 'لا توجد'}<br>
                📍 آخر عنوان: {last_v.get('address','')[:60]}<br>
                ⭐ آخر تقييم: {last_v.get('rating',0) or 'لم يقيم'}
              </div></div>""", unsafe_allow_html=True)
            col_use, col_ignore = st.columns(2)
            with col_use:
                if st.button("✅ استخدم بياناته السابقة", key="use_prev_data", use_container_width=True):
                    st.session_state.prefill = {
                        "name": last_v.get("name",""), "age": last_v.get("age",""),
                        "age_unit": last_v.get("age_unit","سنة"), "phone": last_v.get("phone",""),
                        "address": last_v.get("address",""), "location_link": last_v.get("location_link",""),
                        "doctor_name": last_v.get("doctor_name",""), "branch": last_v.get("branch","La Cite"),
                        "selected_labs_text": last_v.get("selected_labs_text",""),
                        "labs_price_before": last_v.get("labs_price_before",0),
                        "labs_price_after": last_v.get("labs_price_after",0),
                        "transport_fee": last_v.get("transport_fee",100),
                        "visit_time": "", "notes": "",
                    }
                    st.rerun()
            with col_ignore:
                if "ignore_dup_warning" not in st.session_state: st.session_state.ignore_dup_warning = False
                if st.button("➕ متابعة كزيارة جديدة", key="ignore_dup", use_container_width=True):
                    st.session_state.ignore_dup_warning = True; st.rerun()
        else:
            st.markdown('<div style="background:#D5F5E3;border:1.5px solid #27AE60;border-radius:10px;padding:8px 14px;margin:6px 0;font-size:13px;color:#1E8449;">🆕 عميل جديد — لم يسبق له زيارة</div>', unsafe_allow_html=True)
    DOCTOR_LIST = get_doctor_names() + ["أخرى..."]
    saved_doc   = pf.get("doctor_name","")
    doc_index   = DOCTOR_LIST.index(saved_doc) if saved_doc in DOCTOR_LIST else (len(DOCTOR_LIST)-1 if saved_doc else 0)
    doc_col1, doc_col2 = st.columns([3,1])
    with doc_col1:
        doc_select = st.selectbox("👨‍⚕️ الدكتور القائم بالزيارة", DOCTOR_LIST, index=doc_index)
        if doc_select == "أخرى...":
            doctor_name = st.text_input("اكتب اسم الدكتور", value=saved_doc if saved_doc not in DOCTOR_LIST else "", placeholder="اسم الدكتور...")
        else:
            doctor_name = doc_select
    with doc_col2:
        st.markdown('<div style="margin-top:28px"></div>', unsafe_allow_html=True)
        dc1,dc2 = st.columns(2)
        with dc1:
            visit_date_temp = date.today()
            if pf.get("visit_date"):
                try: visit_date_temp = datetime.strptime(pf["visit_date"],"%Y-%m-%d").date()
                except: pass
            workload = get_doctor_workload(doctor_name, visit_date_temp)
        if workload > 0:
            wl_color = "#E74C3C" if workload >= 8 else "#F39C12" if workload >= 5 else "#27AE60"
            st.markdown(f'<div style="background:{wl_color};color:#fff;border-radius:8px;padding:6px 10px;font-size:11px;font-weight:700;text-align:center;">{workload} زيارة اليوم</div>', unsafe_allow_html=True)
    branch    = st.selectbox("🏥 الفرع", ["La Cite","Diamond"], index=0 if pf.get("branch","La Cite")=="La Cite" else 1)
    cur_status = pf.get("status","مجدولة")
    if cur_status not in STATUS_OPTIONS: cur_status = "مجدولة"
    status = st.selectbox("🔖 حالة الزيارة", STATUS_OPTIONS, index=STATUS_OPTIONS.index(cur_status))
    dc1,dc2 = st.columns(2)
    with dc1:
        default_date = date.today()
        if pf.get("visit_date"):
            try: default_date = datetime.strptime(pf["visit_date"],"%Y-%m-%d").date()
            except: pass
        visit_date = st.date_input("📅 تاريخ الزيارة *", value=default_date)
        if visit_date < date.today() and not is_edit:
            st.warning("⚠️ التاريخ في الماضي — هل أنت متأكد؟")
    with dc2:
        st.markdown("🕐 وقت الزيارة")
        tc1,tc2,tc3 = st.columns([2,2,3])
        old_t = pf.get("visit_time",""); ph,pm,pa = 12,0,"PM"
        if old_t:
            m = re_module.match(r'(\d{1,2}):(\d{2})\s*(AM|PM)', old_t, re_module.IGNORECASE)
            if m: ph,pm,pa = int(m.group(1)),int(m.group(2)),m.group(3).upper()
        with tc1: hour   = st.selectbox("ساعة", list(range(1,13)), index=ph-1 if 1<=ph<=12 else 11, key="hr_sel")
        with tc2: minute = st.selectbox("دقيقة", [0,15,30,45], index=[0,15,30,45].index(pm) if pm in [0,15,30,45] else 0, key="mn_sel")
        with tc3: ampm   = st.radio("", ["AM","PM"], index=0 if pa=="AM" else 1, horizontal=True, key="ap_sel")
    visit_time = f"{hour}:{minute:02d} {ampm}"
    st.markdown("---")
    st.markdown('<div class="section-title">📍 العنوان</div>', unsafe_allow_html=True)
    address       = st.text_area("العنوان بالتفصيل *", value=pf.get("address",""), placeholder="المحافظة - المدينة - الشارع - رقم المبنى - الدور - الشقة...", height=90)
    location_link = st.text_input("🗺️ رابط الموقع (Google Maps)", value=pf.get("location_link",""))
    st.markdown("---")
    vid_key     = pf.get("id","new_visit")
    labs_ss_key = f"added_labs_{vid_key}"
    if labs_ss_key not in st.session_state:
        if pf.get("selected_labs_text",""):
            st.session_state[labs_ss_key] = [l.strip() for l in pf["selected_labs_text"].splitlines() if l.strip()]
        else:
            st.session_state[labs_ss_key] = []
    st.markdown('<div class="section-title">⚡ Quick Panels</div>', unsafe_allow_html=True)
    st.caption("اضغط على panel لإضافة تحاليله فوراً — التحاليل المكررة لن تُضاف")
    pcols = st.columns(4)
    for i,panel in enumerate(QUICK_PANELS):
        with pcols[i%4]:
            if st.button(panel["name"], key=f"pnl_{vid_key}_{i}", use_container_width=True):
                existing = [e.split(" — ")[0].strip() for e in st.session_state[labs_ss_key]]
                for tn in panel["tests"]:
                    if tn not in existing: st.session_state[labs_ss_key].append(tn)
                st.rerun()
    with st.expander("👁️ شاهد محتوى الـ Panels"):
        for panel in QUICK_PANELS:
            st.markdown(f'<div style="font-size:12px;margin-bottom:6px"><b style="color:#FF6B00">{panel["name"]}</b> — {" • ".join(panel["tests"])}</div>', unsafe_allow_html=True)
    st.markdown("---")
    if ALL_LABS:
        st.markdown('<div class="section-title">📋 إضافة تحليل من قائمة الأسعار</div>', unsafe_allow_html=True)
        lab_options = [f"{lab['name']} — {lab['price']} جنيه" for lab in ALL_LABS]
        sel_lab     = st.selectbox("اختر التحليل", lab_options, key=f"lps_{vid_key}")
        if st.button("➕ أضف من القائمة", key=f"alp_{vid_key}", use_container_width=True):
            if sel_lab not in st.session_state[labs_ss_key]: st.session_state[labs_ss_key].append(sel_lab)
            st.rerun()
    else:
        st.warning("قائمة الأسعار غير متاحة حالياً")
    st.markdown("---")
    st.markdown('<div class="section-title">🧪 التحاليل المضافة</div>', unsafe_allow_html=True)
    if st.session_state[labs_ss_key]:
        auto_total = sum(int(m.group(1)) for e in st.session_state[labs_ss_key] for m in [re_module.search(r'(\d+)\s*جنيه', e)] if m)
        st.markdown(f'<div style="font-size:12px;color:#FF6B00;font-weight:700;margin-bottom:8px">✅ {len(st.session_state[labs_ss_key])} تحليل{"  —  إجمالي: "+f"{auto_total:,} جنيه" if auto_total else ""}</div>', unsafe_allow_html=True)
        to_remove = None
        for i,entry in enumerate(st.session_state[labs_ss_key]):
            ca,cb = st.columns([10,1])
            with ca: st.markdown(f'<div style="font-size:13px;padding:4px 0;border-bottom:1px solid #f5f5f5;color:#333">🔹 {entry}</div>', unsafe_allow_html=True)
            with cb:
                if st.button("✕", key=f"del_{vid_key}_{i}", help="احذف"): to_remove = i
        if to_remove is not None: st.session_state[labs_ss_key].pop(to_remove); st.rerun()
        if st.button("🗑️ مسح الكل", key=f"clr_{vid_key}"):
            st.session_state[labs_ss_key] = []; st.rerun()
    else:
        st.markdown('<div style="color:#aaa;font-size:13px;padding:8px 0">لا توجد تحاليل — اختر panel أو أضف يدوياً</div>', unsafe_allow_html=True)
    cm1,cm2 = st.columns([8,2])
    with cm1: manual_entry = st.text_input("أضف تحليل يدوياً", placeholder="CBC — 400 جنيه", key=f"man_{vid_key}")
    with cm2:
        st.markdown('<div style="margin-top:28px"></div>', unsafe_allow_html=True)
        if st.button("➕ أضف", key=f"manb_{vid_key}", use_container_width=True):
            if manual_entry.strip(): st.session_state[labs_ss_key].append(manual_entry.strip()); st.rerun()
    selected_labs_text = "\n".join(st.session_state[labs_ss_key])
    selected_labs      = st.session_state[labs_ss_key][:]
    if st.session_state[labs_ss_key] and LABS_PRICE_LOOKUP:
        sugg_all, bundle_sugg = get_smart_suggestions(st.session_state[labs_ss_key], LABS_PRICE_LOOKUP, phone=phone if phone else None)
        if sugg_all or bundle_sugg:
            st.markdown('<div style="background:linear-gradient(135deg,#1a1a2e,#16213e);border-radius:14px;padding:12px 16px;margin:12px 0;"><div style="color:#FF6B00;font-weight:800;font-size:14px;margin-bottom:4px;">🤖 اقتراحات ذكية</div></div>', unsafe_allow_html=True)
            for b in bundle_sugg:
                st.markdown(f'<div style="background:#FFF9E6;border:2px solid #F39C12;border-radius:10px;padding:10px 14px;margin-bottom:8px;"><b style="color:#E67E22;">💰 عرض توفير</b><br><span style="font-size:12px;color:#333;">{b["note"]}</span></div>', unsafe_allow_html=True)
                if b["action"] == "replace":
                    bc1, bc2 = st.columns(2)
                    with bc1:
                        if st.button(f'🔄 استبدل بـ {b["name"]}', key=f'bundle_rep_{b["name"]}_{vid_key}', use_container_width=True):
                            st.session_state[labs_ss_key] = [e for e in st.session_state[labs_ss_key] if b["remove"] not in e]
                            price = LABS_PRICE_LOOKUP.get(b["name"], b["price"])
                            entry = f'{b["name"]} — {price} جنيه' if price else b["name"]
                            st.session_state[labs_ss_key].append(entry); st.rerun()
                    with bc2: st.button("❌ تجاهل العرض", key=f'bundle_ign_{b["name"]}_{vid_key}', use_container_width=True)
                elif b["action"] == "add_discounted":
                    if st.button(f'➕ أضف فيتامين د ثاني بـ {b["price"]} جنيه', key=f'bundle_add_{vid_key}', use_container_width=True):
                        st.session_state[labs_ss_key].append(f'Vitamin D3(25 Hydroxy Cholecal.) — {b["price"]} جنيه (سعر خاص - الفردين)'); st.rerun()
            if sugg_all:
                shown_reasons = set()
                for lab_name, info in sugg_all.items():
                    if info["reason"] not in shown_reasons:
                        shown_reasons.add(info["reason"])
                        icon = "🔬" if info["type"]=="panel" else "🕐" if info["type"]=="history" else "⚕️"
                        st.markdown(f'<div style="font-size:11px;color:#888;margin:8px 0 4px 0;border-right:3px solid #FF6B00;padding-right:8px;">{icon} {info["reason"]}</div>', unsafe_allow_html=True)
                    price_str = f' — {info["price"]} جنيه' if info["price"] else ""
                    sc1, sc2 = st.columns([7, 3])
                    with sc1:
                        st.markdown(f'<div style="background:#F8F9FA;border-radius:8px;padding:7px 12px;font-size:12px;color:#222;border-right:3px solid #27AE60;">🧪 <b>{lab_name}</b>{price_str}</div>', unsafe_allow_html=True)
                    with sc2:
                        if st.button("➕ أضف", key=f'sugg_{lab_name}_{vid_key}', use_container_width=True):
                            existing_names = [e.split(" — ")[0].strip() for e in st.session_state[labs_ss_key]]
                            if lab_name not in existing_names:
                                entry = f'{lab_name} — {info["price"]} جنيه' if info["price"] else lab_name
                                st.session_state[labs_ss_key].append(entry)
                            st.rerun()
                st.markdown("---")
                if st.button("📋 أضف كل الاقتراحات لملاحظات الدكتور", key=f"add_all_notes_{vid_key}", use_container_width=True):
                    notes_text = "💡 اقتراحات النظام:\n"
                    for lab_name, info in sugg_all.items():
                        p = f' ({info["price"]} ج)' if info["price"] else ""
                        notes_text += f'• {lab_name}{p} — {info["reason"]}\n'
                    st.session_state[f"auto_notes_{vid_key}"] = notes_text; st.rerun()
    st.markdown("---")
    st.markdown('<div class="section-title">📌 ملاحظات</div>', unsafe_allow_html=True)
    auto_notes_val = st.session_state.get(f"auto_notes_{vid_key}", "")
    default_notes  = pf.get("notes","")
    if auto_notes_val and auto_notes_val not in default_notes:
        default_notes = (default_notes + "\n" + auto_notes_val).strip()
    notes = st.text_area("ملاحظات خاصة", value=default_notes, height=75)
    st.markdown("---")
    st.markdown('<div class="section-title">💰 الأسعار</div>', unsafe_allow_html=True)
    auto_labs_total = sum(int(m.group(1)) for e in selected_labs for m in [re_module.search(r'(\d+)\s*جنيه', e)] if m)
    pp1,pp2,pp3 = st.columns(3)
    with pp1: labs_price_before = st.number_input("⭐ السعر قبل الخصم", min_value=0, step=10, value=auto_labs_total if auto_labs_total>0 else int(pf.get("labs_price_before",0) or 0))
    with pp2: labs_price_after  = st.number_input("⭐ السعر بعد الخصم",  min_value=0, step=10, value=int(pf.get("labs_price_after",0) or 0))
    with pp3: transport_fee     = st.number_input("⭐ بدل الانتقال",      min_value=0, step=10, value=int(pf.get("transport_fee",100) or 100))
    total_price = labs_price_after + transport_fee
    st.markdown(f"""<div class="price-box">
      <div class="price-row"><span>⭐ السعر قبل الخصم</span><span>{labs_price_before} جنيه</span></div>
      <div class="price-row"><span>⭐ السعر بعد الخصم</span><span>{labs_price_after} جنيه</span></div>
      <div class="price-row"><span>⭐ بدل الانتقال</span><span>{transport_fee} جنيه</span></div>
      <div class="price-total"><span>⭐ الإجمالي</span><span>{total_price} جنيه</span></div>
    </div>""", unsafe_allow_html=True)
    st.markdown('<div class="section-title">💳 حالة الدفع</div>', unsafe_allow_html=True)
    pay_c1, pay_c2 = st.columns(2)
    with pay_c1:
        init_pay_status = pf.get("payment_status","غير مدفوع")
        if init_pay_status not in PAYMENT_STATUS_OPTIONS: init_pay_status = "غير مدفوع"
        init_pay_st_idx = PAYMENT_STATUS_OPTIONS.index(init_pay_status)
        new_pay_status  = st.selectbox("حالة الدفع", PAYMENT_STATUS_OPTIONS, index=init_pay_st_idx, key="new_pay_status")
    with pay_c2:
        init_pay_method = pf.get("payment_method","")
        pay_method_opts = [""] + PAYMENT_METHODS
        init_pm_idx = pay_method_opts.index(init_pay_method) if init_pay_method in pay_method_opts else 0
        new_pay_method  = st.selectbox("طريقة الدفع", pay_method_opts, index=init_pm_idx, key="new_pay_method")
    if new_pay_status in ["مدفوع جزئياً","مدفوع"]:
        paid_amount = st.number_input("المبلغ المدفوع", min_value=0, step=10, value=int(pf.get("paid_amount",0) or 0), key="paid_amount_input")
    else:
        paid_amount = 0
    if st.button("💾 حفظ الزيارة" if not is_edit else "💾 حفظ التعديلات", use_container_width=True):
        if not name or not phone or not address:
            st.error("⚠️ من فضلك املأ الاسم والتليفون والعنوان")
        else:
            record = {
                "id": pf.get("id", uuid_lib.uuid4().hex[:16]),
                "created_at": pf.get("created_at", datetime.now().isoformat()),
                "name": name, "age": age, "age_unit": age_unit,
                "phone": phone, "visit_date": visit_date.isoformat(),
                "visit_time": visit_time, "doctor_name": doctor_name,
                "branch": branch, "address": address, "location_link": location_link,
                "selected_labs_text": selected_labs_text, "notes": notes,
                "labs_price_before": labs_price_before, "labs_price_after": labs_price_after,
                "transport_fee": transport_fee, "total_price": total_price,
                "status": status,
                "payment_status": new_pay_status, "payment_method": new_pay_method,
                "paid_amount": paid_amount, "payment_date": date.today().isoformat() if new_pay_status=="مدفوع" else pf.get("payment_date",""),
                "_user": st.session_state.user_email or "",
            }
            if is_edit: update_visit(record); st.success("✅ تم تحديث الزيارة!")
            else:       insert_visit(record);  st.success("✅ تم حفظ الزيارة!")
            go("detail", visit_id=record["id"])
    if is_edit:
        if st.button("← رجوع بدون حفظ", use_container_width=True):
            go("detail", visit_id=pf.get("id"))

# ══════════════════════════════════════════════════════════════════════════════
# صفحة التفاصيل
# ══════════════════════════════════════════════════════════════════════════════
elif st.session_state.page == "detail":
    vid = st.session_state.selected_id
    v   = fetch_visit_by_id(vid) if vid else None
    if not v:
        st.error("لم يتم العثور على الزيارة"); go("home")
    else:
        lpb   = v.get("labs_price_before",0); lpa = v.get("labs_price_after",0)
        tf    = v.get("transport_fee",0);     tp  = v.get("total_price",0)
        vtime = v.get("visit_time","")
        dt_disp = format_date_ar(v.get("visit_date","")) + (f" — {vtime}" if vtime else "")
        age   = v.get("age",""); au = v.get("age_unit","سنة")
        age_str = f"🎂 {age} {au}" if age else "🎂 غير محدد"
        status = v.get("status","مجدولة")
        sc    = STATUS_COLORS.get(status,"#888"); si = STATUS_ICONS.get(status,"")
        pay_status = v.get("payment_status","غير مدفوع")
        pay_color  = PAYMENT_COLORS.get(pay_status,"#888")
        pay_icon   = PAYMENT_ICONS.get(pay_status,"🔴")

        st.markdown('<div class="section-title">👤 البيانات الشخصية</div>', unsafe_allow_html=True)
        st.markdown(f"""
        <div class="detail-row"><span class="detail-label">👤 الاسم</span><span class="detail-value">{v['name']}</span></div>
        <div class="detail-row"><span class="detail-label">🎂 السن</span><span class="detail-value">{age_str}</span></div>
        <div class="detail-row"><span class="detail-label">📞 التليفون</span><span class="detail-value">{v.get('phone','')}</span></div>
        <div class="detail-row"><span class="detail-label">📅 الموعد</span><span class="detail-value">{dt_disp}</span></div>
        <div class="detail-row"><span class="detail-label">👨‍⚕️ الدكتور</span><span class="detail-value">{v.get('doctor_name','')}</span></div>
        <div class="detail-row"><span class="detail-label">🏥 الفرع</span><span class="detail-value">{v.get('branch','')}</span></div>
        <div class="detail-row"><span class="detail-label">🔖 الحالة</span>
          <span class="detail-value"><span class="status-badge" style="background:{sc}">{si} {status}</span></span></div>
        <div class="detail-row"><span class="detail-label">💳 الدفع</span>
          <span class="detail-value"><span class="pay-badge" style="background:{pay_color}">{pay_icon} {pay_status}</span>
          {'  —  ' + str(v.get('paid_amount',0)) + ' جنيه' if pay_status=='مدفوع جزئياً' else ''}
          {'  —  ' + v.get('payment_method','') if v.get('payment_method') else ''}</span></div>
        """, unsafe_allow_html=True)

        tag_auto  = get_client_tag(v.get("phone",""))
        tag_color = get_client_tag_color(tag_auto)
        churn     = get_churn_risk(v.get("phone",""))
        all_visits_c = fetch_client_history(v.get("phone",""))
        done_count_d = sum(1 for hv in all_visits_c if hv.get("status")=="تمت")
        cur_rating   = int(v.get("rating", 0) or 0)
        stars_html   = f'<span style="font-size:16px;">{"⭐"*cur_rating}</span>' if cur_rating else ""
        st.markdown(
            f'<div style="display:flex;gap:10px;align-items:center;margin-bottom:12px;flex-wrap:wrap;">'
            f'<span style="background:{tag_color};color:#fff;border-radius:20px;padding:5px 14px;font-weight:800;font-size:13px;">{tag_auto}</span>'
            f'<span style="color:#888;font-size:12px;">زيارات مكتملة: <b style="color:#27AE60">{done_count_d}</b></span>'
            f'<span style="background:#f0f0f0;color:#555;border-radius:8px;padding:3px 10px;font-size:12px;">📊 {churn}</span>'
            f'{stars_html}</div>', unsafe_allow_html=True)
        tag_options = ["🆕 عميل جديد","⭐ عميل منتظم","🌟 عميل متكرر","👑 VIP","🏢 Corporate"]
        saved_tag   = v.get("tag","") or tag_auto
        if saved_tag not in tag_options: saved_tag = tag_auto
        new_tag = st.selectbox("🏷️ تصنيف العميل", tag_options,
                               index=tag_options.index(saved_tag) if saved_tag in tag_options else 0,
                               key=f"tag_sel_{v['id']}")
        if new_tag != v.get("tag",""): update_tag(v["id"], new_tag)
        st.markdown("---")
        st.markdown("**⚡ تغيير الحالة السريع:**")
        qs_cols = st.columns(4)
        for col, sn in zip(qs_cols, STATUS_OPTIONS):
            with col:
                prefix = "✓ " if sn == status else ""
                if st.button(f"{prefix}{STATUS_ICONS[sn]} {sn}", key=f"qs_{v['id']}_{sn}", use_container_width=True):
                    update_status_only(v["id"], sn); st.rerun()
        st.markdown("---")

        st.markdown('<div class="section-title">📍 العنوان</div>', unsafe_allow_html=True)
        st.write(v.get("address",""))
        if v.get("location_link"):
            st.markdown(f'<a href="{v["location_link"]}" target="_blank" style="color:#FF6B00;font-weight:700;font-size:13px;">🗺️ فتح الموقع على الخريطة</a>', unsafe_allow_html=True)
        st.markdown("---")

        st.markdown('<div class="section-title">🧪 التحاليل المطلوبة</div>', unsafe_allow_html=True)
        lt = v.get("selected_labs_text","")
        if lt.strip():
            labs_count = 0
            for lab in lt.splitlines():
                if lab.strip():
                    st.markdown(f'<div style="font-size:13px;padding:5px 0;border-bottom:1px solid #f5f5f5;color:#333;">🔹 {lab.strip()}</div>', unsafe_allow_html=True)
                    labs_count += 1
            st.markdown(f'<div style="font-size:12px;color:#FF6B00;font-weight:700;margin-top:8px;">إجمالي: {labs_count} تحليل</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div style="color:#aaa;font-size:13px;">لا توجد تحاليل مسجلة</div>', unsafe_allow_html=True)

        if v.get("notes"):
            st.markdown("---")
            st.markdown('<div class="section-title">📌 ملاحظات</div>', unsafe_allow_html=True)
            st.write(v.get("notes",""))
        st.markdown("---")

        st.markdown(f"""<div class="price-box">
          <div class="price-row"><span>⭐ السعر قبل الخصم</span><span>{format_money(lpb)}</span></div>
          <div class="price-row"><span>⭐ السعر بعد الخصم</span><span>{format_money(lpa)}</span></div>
          <div class="price-row"><span>🚗 بدل الانتقال</span><span>{format_money(tf)}</span></div>
          <div class="price-total"><span>💵 الإجمالي</span><span>{format_money(tp)}</span></div>
        </div>""", unsafe_allow_html=True)

        with st.expander("💳 تحديث حالة الدفع", expanded=(pay_status=="غير مدفوع")):
            pc1, pc2 = st.columns(2)
            with pc1:
                pi = PAYMENT_STATUS_OPTIONS.index(pay_status) if pay_status in PAYMENT_STATUS_OPTIONS else 0
                new_pay_st = st.selectbox("حالة الدفع", PAYMENT_STATUS_OPTIONS, index=pi, key="det_pay_st")
            with pc2:
                pm_opts = [""] + PAYMENT_METHODS
                cur_pm  = v.get("payment_method","")
                pm_idx  = pm_opts.index(cur_pm) if cur_pm in pm_opts else 0
                new_pay_method = st.selectbox("طريقة الدفع", pm_opts, index=pm_idx, key="det_pay_method")
            if new_pay_st in ["مدفوع جزئياً","مدفوع"]:
                new_paid_amount = st.number_input("المبلغ المدفوع (جنيه)", min_value=0, step=10,
                                                   value=int(v.get("paid_amount",0) or 0), key="det_paid_amount")
            else:
                new_paid_amount = 0
            if st.button("💾 حفظ تحديث الدفع", key="save_payment", use_container_width=True):
                pay_date = date.today().isoformat() if new_pay_st in ["مدفوع جزئياً","مدفوع"] else ""
                update_payment(v["id"], new_pay_st, new_pay_method, new_paid_amount, pay_date,
                               user_email=st.session_state.user_email or "")
                st.success("✅ تم تحديث حالة الدفع!"); st.rerun()

        st.markdown("---")

        # ── تقييم الزيارة ──
        st.markdown('<div class="section-title">📊 تقييم الزيارة</div>', unsafe_allow_html=True)
        if status == "تمت":
            rating_cols = st.columns(6)
            with rating_cols[0]:
                st.markdown('<div style="font-size:13px;color:#888;padding-top:8px;">التقييم:</div>', unsafe_allow_html=True)
            for star_n in range(1,6):
                with rating_cols[star_n]:
                    star_icon = "⭐" if star_n <= cur_rating else "☆"
                    if st.button(f"{star_icon}{star_n}", key=f"star_{v['id']}_{star_n}", use_container_width=True):
                        update_rating(v["id"], star_n); st.rerun()
            if cur_rating:
                st.markdown(f'<div style="text-align:center;font-size:13px;color:#FF6B00;font-weight:700;margin-top:4px;">التقييم الحالي: {"⭐"*cur_rating} ({cur_rating}/5)</div>', unsafe_allow_html=True)
            rating_msg = (
                f"🟠 *Orange Lab Home Visit Management System*\n━━━━━━━━━━━━━━\n"
                f"أهلاً {v.get('name','')} 🌟\n"
                "شكراً لاختيارك معمل أورانج لاب للزيارة المنزلية!\n\n"
                "نتمنى أن الخدمة كانت على مستوى توقعاتك 🧡\n\n"
                "⭐ *قيّم خدمتنا من 1 إلى 5:*\n"
                "  1️⃣ - ضعيف\n  2️⃣ - مقبول\n  3️⃣ - جيد\n  4️⃣ - جيد جداً\n  5️⃣ - ممتاز\n\n"
                "رأيك يهمنا لتحسين خدماتنا 💛\n━━━━━━━━━━━━━━\n*معمل أورانج لاب*"
            )
            st.markdown(f'<a href="{whatsapp_link(rating_msg, v.get("phone"))}" target="_blank" class="wa-btn" style="background:#9B59B6;margin-top:8px;">📊 إرسال طلب تقييم على واتساب</a>', unsafe_allow_html=True)
        else:
            st.info("التقييم متاح فقط بعد اكتمال الزيارة (حالة: تمت)")
        st.markdown("---")

        # ── واتساب ──
        st.markdown('<div class="section-title">📲 رسائل واتساب</div>', unsafe_allow_html=True)
        msg_client   = make_whatsapp_msg(v, "client")
        msg_internal = make_whatsapp_msg(v, "internal")
        msg_group    = make_whatsapp_msg(v, "group")
        phone_clean  = v.get("phone","")
        st.markdown(f'<a href="{whatsapp_link(msg_client, phone_clean)}" target="_blank" class="wa-btn wa-client">📱 رسالة للعميل</a>', unsafe_allow_html=True)
        st.markdown(f'<a href="{whatsapp_link(msg_internal)}" target="_blank" class="wa-btn wa-share">🔗 مشاركة داخلية</a>', unsafe_allow_html=True)
        st.markdown(f'<a href="{whatsapp_link(msg_group)}" target="_blank" class="wa-btn wa-group">👥 رسالة الجروب</a>', unsafe_allow_html=True)
        st.markdown(f'<a href="tel:{phone_clean}" class="wa-btn" style="background:#3498DB;">📞 اتصال مباشر</a>', unsafe_allow_html=True)
        st.markdown("---")

        # ── تاريخ العميل ──
        prev = fetch_client_history(v.get("phone",""), exclude_id=v["id"])
        if prev:
            st.markdown('<div class="section-title">📋 تاريخ العميل</div>', unsafe_allow_html=True)
            for old_v in prev[:4]:
                old_date = format_date_ar(old_v.get("visit_date",""))
                old_stat = old_v.get("status","")
                old_sc   = STATUS_COLORS.get(old_stat,"#888")
                old_pay  = old_v.get("payment_status","غير مدفوع")
                old_pc   = PAYMENT_COLORS.get(old_pay,"#888")
                st.markdown(f'<div class="history-card">📅 {old_date} — <span style="background:{old_sc};color:#fff;border-radius:6px;padding:1px 8px;font-size:11px;">{old_stat}</span>'
                            f' <span style="background:{old_pc};color:#fff;border-radius:6px;padding:1px 8px;font-size:11px;">{PAYMENT_ICONS.get(old_pay,"")} {old_pay}</span>'
                            f'<br>💰 {format_money(old_v.get("total_price",0))} — 🧪 {len((old_v.get("selected_labs_text","") or "").splitlines())} تحليل</div>',
                            unsafe_allow_html=True)
                if st.button(f"📂 فتح", key=f"hist_{old_v['id']}", use_container_width=True):
                    go("detail", visit_id=old_v["id"])
            if st.button("👤 بروفايل العميل الكامل", key="view_full_profile", use_container_width=True):
                go("client_profile", client_phone=v.get("phone",""))
            st.markdown("---")

        # ── متابعات مرتبطة بالزيارة ──
        visit_followups = fetch_follow_ups({"phone": v.get("phone","")})
        active_fu = [f for f in visit_followups if not f.get("done")]
        if active_fu:
            st.markdown('<div class="section-title">⏰ متابعات معلقة</div>', unsafe_allow_html=True)
            for fu in active_fu[:3]:
                fu_date  = format_date_ar(fu.get("follow_up_date",""))
                is_over  = fu.get("follow_up_date","") < date.today().isoformat()
                card_cls = "fu-card fu-overdue" if is_over else "fu-card"
                icon     = "🔴" if is_over else "⏰"
                st.markdown(f'<div class="{card_cls}">{icon} <b>{fu_date}</b> — {fu.get("reason","")}</div>', unsafe_allow_html=True)
            st.markdown("---")

        # ── إنشاء متابعة جديدة ──
        with st.expander("➕ إضافة متابعة جديدة"):
            fu_date_input = st.date_input("تاريخ المتابعة", value=date.today() + timedelta(days=7), key="fu_date_input")
            fu_reason     = st.text_input("سبب المتابعة", placeholder="مثال: متابعة نتائج TSH بعد أسبوع", key="fu_reason")
            if st.button("💾 حفظ المتابعة", key="save_fu", use_container_width=True):
                if fu_date_input and fu_reason.strip():
                    insert_follow_up(
                        visit_id=v["id"], client_name=v.get("name",""),
                        client_phone=v.get("phone",""), follow_up_date=fu_date_input.isoformat(),
                        reason=fu_reason.strip(), created_by=st.session_state.user_email or ""
                    )
                    st.success("✅ تم إضافة المتابعة!"); st.rerun()
                else: st.error("أدخل السبب والتاريخ")

        st.markdown("---")

        # ── تحميل ورقة الزيارة ──
        print_html = generate_visit_print_html(v)
        st.download_button(label="⬇️ تحميل ورقة الزيارة (HTML للطباعة)",
            data=print_html.encode("utf-8"),
            file_name=f"زيارة_{v['name']}_{v.get('visit_date','')}.html",
            mime="text/html", use_container_width=True)
        st.markdown("---")

        ec1,ec2 = st.columns(2)
        with ec1:
            if st.button("✏️ تعديل", use_container_width=True): go("new", prefill={**v, "_edit":True})
        with ec2:
            if st.button("🗑️ حذف", use_container_width=True): st.session_state["confirm_delete"] = True
        if st.session_state.get("confirm_delete"):
            st.warning("⚠️ هل أنت متأكد من الحذف؟ (الزيارة ستُخفى وليس حذف نهائي)")
            dc1,dc2 = st.columns(2)
            with dc1:
                if st.button("✅ نعم، احذف", use_container_width=True):
                    soft_delete_visit(vid, user_email=st.session_state.user_email or "")
                    st.session_state["confirm_delete"] = False; go("home")
            with dc2:
                if st.button("❌ إلغاء", use_container_width=True):
                    st.session_state["confirm_delete"] = False; st.rerun()

        st.markdown(f'<div class="repeat-banner">🔄 هتروح لـ {v["name"]} مرة تانية؟</div>', unsafe_allow_html=True)
        if st.button(f"➕ زيارة جديدة لـ {v['name']}", use_container_width=True):
            go("new", prefill={
                "name":v["name"],"age":v.get("age",""),"age_unit":v.get("age_unit","سنة"),
                "phone":v.get("phone",""),"address":v.get("address",""),
                "location_link":v.get("location_link",""),"doctor_name":v.get("doctor_name",""),
                "branch":v.get("branch","La Cite"),"selected_labs_text":"",
                "visit_time":"","notes":"","labs_price_before":0,"labs_price_after":0,"transport_fee":100,
            })
        if st.button("← رجوع للقائمة", use_container_width=True): go("home")

# ══════════════════════════════════════════════════════════════════════════════
# صفحة التقارير
# ══════════════════════════════════════════════════════════════════════════════
elif st.session_state.page == "reports":
    if st.session_state.user_type not in ["admin","diamond","lacite"]:
        st.error("غير مصرح."); st.stop()
    st.markdown("### 📈 التقارير")
    import plotly.graph_objects as go_plotly
    import plotly.express as px
    rf1,rf2,rf3 = st.columns(3)
    with rf1:
        rep_year  = st.selectbox("السنة", list(range(datetime.now().year, datetime.now().year-4, -1)), index=0, key="rep_year")
    with rf2:
        rep_month = st.selectbox("الشهر (0=كل السنة)", list(range(0,13)), index=0, key="rep_month",
                                  format_func=lambda x: "كل السنة" if x==0 else MONTHS_AR[x-1])
    with rf3:
        if st.session_state.user_type == "diamond":   rep_branch = "Diamond"
        elif st.session_state.user_type == "lacite":  rep_branch = "La Cite"
        else:
            rep_branch_sel = st.selectbox("الفرع", ["كل الفروع","La Cite","Diamond"], key="rep_branch")
            rep_branch = None if rep_branch_sel == "كل الفروع" else rep_branch_sel
    rf = {}
    if rep_branch: rf["branch"] = rep_branch
    if rep_month > 0: rf["month"] = rep_month; rf["year"] = rep_year
    else:             rf["year"]  = rep_year
    if rf.get("year") and not rf.get("month"):
        year_str = str(rf["year"])
        conn = get_connection()
        if rf.get("branch"):
            rows = conn.execute(
                "SELECT * FROM visits WHERE strftime('%Y',visit_date)=? AND branch=? AND deleted_at IS NULL",
                (year_str, rf["branch"])
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM visits WHERE strftime('%Y',visit_date)=? AND deleted_at IS NULL", (year_str,)
            ).fetchall()
        visits_r = [dict(r) for r in rows]
    else:
        visits_r, _ = fetch_visits(rf, page=None, page_size=None)
    if not visits_r:
        st.info("لا توجد بيانات للفترة المحددة.")
    else:
        df_r = pd.DataFrame(visits_r)
        total_r = len(df_r)
        rev_r   = df_r["labs_price_after"].sum()
        done_r  = (df_r["status"]=="تمت").sum()
        cancelled_r = (df_r["status"]=="ملغية").sum()
        unpaid_r   = (df_r.get("payment_status","غير مدفوع")=="غير مدفوع").sum() if "payment_status" in df_r.columns else 0
        paid_r     = (df_r.get("payment_status","")=="مدفوع").sum() if "payment_status" in df_r.columns else 0
        collect_pct = round(paid_r/max(done_r,1)*100) if done_r > 0 else 0
        st.markdown(f"""<div class="stat-grid">
          <div class="stat-box"><div class="stat-num">{total_r}</div><div class="stat-label">إجمالي الزيارات</div></div>
          <div class="stat-box"><div class="stat-num" style="color:#27AE60">{done_r}</div><div class="stat-label">تمت ✅</div></div>
          <div class="stat-box"><div class="stat-num" style="color:#E74C3C">{cancelled_r}</div><div class="stat-label">ملغية ❌</div></div>
          <div class="stat-box"><div class="stat-num" style="font-size:15px">{rev_r:,.0f}</div><div class="stat-label">إيراد التحاليل</div></div>
          <div class="stat-box"><div class="stat-num" style="color:#E74C3C">{unpaid_r}</div><div class="stat-label">غير مدفوع 🔴</div></div>
          <div class="stat-box"><div class="stat-num" style="color:#27AE60">{collect_pct}%</div><div class="stat-label">نسبة التحصيل</div></div>
        </div>""", unsafe_allow_html=True)
        try:
            df_r["visit_date_dt"] = pd.to_datetime(df_r["visit_date"], errors="coerce")
            df_r["week"]          = df_r["visit_date_dt"].dt.to_period("W").astype(str)
            df_r["day"]           = df_r["visit_date_dt"].dt.day_name()
            weekly = df_r[df_r["status"]!="ملغية"].groupby("week").agg(count=("id","count"), rev=("labs_price_after","sum")).reset_index()
            if not weekly.empty:
                fig = go_plotly.Figure()
                fig.add_bar(x=weekly["week"], y=weekly["count"], name="عدد الزيارات", marker_color="#3498DB")
                fig.update_layout(title="📊 الزيارات أسبوعياً", xaxis_title="الأسبوع", yaxis_title="العدد",
                                  font=dict(family="Cairo"), height=300, margin=dict(t=40,b=20,l=20,r=20))
                st.plotly_chart(fig, use_container_width=True)
        except Exception as e:
            st.caption(f"تعذّر رسم المخطط: {e}")
        try:
            doc_df = df_r[df_r["status"]!="ملغية"].groupby("doctor_name").agg(count=("id","count"), rev=("labs_price_after","sum")).reset_index().sort_values("count",ascending=False)
            if not doc_df.empty:
                fig2 = px.bar(doc_df, x="doctor_name", y="count", title="👨‍⚕️ الزيارات بالدكتور",
                              labels={"doctor_name":"الدكتور","count":"عدد الزيارات"}, color="count",
                              color_continuous_scale="Oranges")
                fig2.update_layout(font=dict(family="Cairo"), height=280, margin=dict(t=40,b=20,l=20,r=20))
                st.plotly_chart(fig2, use_container_width=True)
        except: pass
        try:
            if "payment_status" in df_r.columns:
                pay_df = df_r[df_r["status"]!="ملغية"]["payment_status"].value_counts().reset_index()
                pay_df.columns = ["status","count"]
                pay_colors_list = [PAYMENT_COLORS.get(s,"#888") for s in pay_df["status"]]
                fig3 = go_plotly.Figure(go_plotly.Pie(
                    labels=pay_df["status"], values=pay_df["count"],
                    marker_colors=pay_colors_list, hole=0.4))
                fig3.update_layout(title="💳 توزيع حالات الدفع", font=dict(family="Cairo"),
                                   height=280, margin=dict(t=40,b=20,l=20,r=20))
                st.plotly_chart(fig3, use_container_width=True)
        except: pass
        try:
            status_df = df_r["status"].value_counts().reset_index()
            status_df.columns = ["status","count"]
            s_colors = [STATUS_COLORS.get(s,"#888") for s in status_df["status"]]
            fig4 = go_plotly.Figure(go_plotly.Pie(labels=status_df["status"], values=status_df["count"],
                marker_colors=s_colors, hole=0.4))
            fig4.update_layout(title="🔖 توزيع الحالات", font=dict(family="Cairo"),
                               height=280, margin=dict(t=40,b=20,l=20,r=20))
            st.plotly_chart(fig4, use_container_width=True)
        except: pass
        # ── إيراد يومي ──
        try:
            daily = {}
            for v in visits_r:
                if v.get("status")=="ملغية": continue
                d = v.get("visit_date","")
                daily[d] = daily.get(d,0) + v.get("total_price",0)
            if daily:
                st.markdown("#### 📊 الإيراد اليومي")
                df_daily = pd.DataFrame(sorted(daily.items()), columns=["التاريخ","الإيراد"])
                st.bar_chart(df_daily.set_index("التاريخ"))
        except: pass
        # ── توزيع على الأطباء ──
        try:
            doc_counts = {}
            for v in visits_r:
                doc = v.get("doctor_name","غير محدد") or "غير محدد"
                doc_counts[doc] = doc_counts.get(doc,0)+1
            if doc_counts:
                st.markdown("#### 👨‍⚕️ توزيع الزيارات على الأطباء")
                df_dc = pd.DataFrame(doc_counts.items(), columns=["الدكتور","عدد الزيارات"])
                st.bar_chart(df_dc.set_index("الدكتور"))
        except: pass
        # ── جدول ملخص الأطباء ──
        try:
            summary = {}
            for v in visits_r:
                doc = v.get("doctor_name","غير محدد") or "غير محدد"
                if doc not in summary:
                    summary[doc] = {"count":0,"before":0,"after":0,"transport":0,"total":0,"done":0,"cancelled":0}
                summary[doc]["count"]    += 1
                summary[doc]["before"]   += v.get("labs_price_before",0)
                summary[doc]["after"]    += v.get("labs_price_after",0)
                summary[doc]["transport"]+= v.get("transport_fee",0)
                if v.get("status")!="ملغية": summary[doc]["total"] += v.get("total_price",0)
                if v.get("status")=="تمت":   summary[doc]["done"]  += 1
                if v.get("status")=="ملغية": summary[doc]["cancelled"] += 1
            if summary:
                st.markdown("#### 📋 ملخص تفصيلي بالأطباء")
                df_sum = pd.DataFrame(summary).T
                df_sum["الدكتور"] = df_sum.index
                df_sum = df_sum[["الدكتور","count","done","cancelled","before","after","transport","total"]]
                df_sum.columns = ["الدكتور","الزيارات","تمت","ملغية","قبل الخصم","بعد الخصم","الانتقال","الإجمالي"]
                df_sum = df_sum.sort_values("الزيارات",ascending=False)
                tc=df_sum["الزيارات"].sum(); tt=df_sum["الإجمالي"].sum()
                st.markdown(f"**إجمالي:** {tc} زيارة &nbsp;|&nbsp; **الإيراد الكلي:** {format_money(tt)}")
                st.dataframe(df_sum.style.format({"قبل الخصم":"{:,.0f} ج","بعد الخصم":"{:,.0f} ج","الانتقال":"{:,.0f} ج","الإجمالي":"{:,.0f} ج"}), use_container_width=True)
                # تقرير HTML للطباعة
                period_label_r = f"{MONTHS_AR[rep_month-1]} {rep_year}" if rep_month > 0 else str(rep_year)
                branch_title_r = f" - فرع {rep_branch}" if rep_branch else ""
                rows_html = ""
                for _,row in df_sum.iterrows():
                    rows_html += f"<tr><td>{row['الدكتور']}</td><td>{row['الزيارات']}</td><td style='color:#27AE60;font-weight:700'>{row['تمت']}</td><td style='color:#E74C3C'>{row['ملغية']}</td><td>{row['قبل الخصم']:,.0f} ج</td><td>{row['بعد الخصم']:,.0f} ج</td><td>{row['الانتقال']:,.0f} ج</td><td><b>{row['الإجمالي']:,.0f} ج</b></td></tr>"
                printable_html = f"""<div id="printable-report" style="direction:rtl;font-family:'Cairo',sans-serif;padding:20px;background:white;color:black;">
                    <h1 style="color:#FF6B00;text-align:center;">Orange Lab Home Visit Management System</h1>
                    <h2 style="text-align:center;">تقرير زيارات {period_label_r}{branch_title_r}</h2>
                    <table border="1" cellpadding="8" cellspacing="0" style="width:100%;border-collapse:collapse;margin-top:20px;">
                    <thead><tr style="background:#FF6B00;color:white;"><th>الدكتور</th><th>الزيارات</th><th>تمت</th><th>ملغية</th><th>قبل الخصم</th><th>بعد الخصم</th><th>الانتقال</th><th>الإجمالي</th></tr></thead>
                    <tbody>{rows_html}
                    <tr style="background:#f5f5f5;font-weight:bold;"><td>الإجمالي</td><td>{df_sum['الزيارات'].sum()}</td><td style="color:#27AE60">{df_sum['تمت'].sum()}</td><td style="color:#E74C3C">{df_sum['ملغية'].sum()}</td><td>{df_sum['قبل الخصم'].sum():,.0f} ج</td><td>{df_sum['بعد الخصم'].sum():,.0f} ج</td><td>{df_sum['الانتقال'].sum():,.0f} ج</td><td>{df_sum['الإجمالي'].sum():,.0f} ج</td></tr>
                    </tbody></table>
                    <p style="text-align:center;margin-top:30px;">تم إنشاؤه بواسطة Orange Lab Home Visit Management System</p>
                </div>"""
                import streamlit.components.v1 as components
                components.html(printable_html, height=500, scrolling=True)
                if st.session_state.user_type == "admin":
                    csv = df_sum.to_csv(index=False).encode("utf-8-sig")
                    st.download_button("📥 تحميل CSV", data=csv,
                                       file_name=f"تقرير_زيارات_{period_label_r}.csv", mime="text/csv", key="dl_csv_rep")
        except Exception as e:
            st.caption(f"تعذّر عرض جدول الملخص: {e}")
        st.markdown("---")
        st.markdown("### 📤 تصدير التقرير")
        exp_c1, exp_c2 = st.columns(2)
        with exp_c1: exp_from_r = st.date_input("من تاريخ", key="rep_exp_from", value=None)
        with exp_c2: exp_to_r   = st.date_input("إلى تاريخ", key="rep_exp_to",   value=date.today())
        if st.button("📤 تصدير Excel", use_container_width=True, key="export_rep"):
            if not exp_from_r or not exp_to_r:
                st.error("اختر التواريخ")
            else:
                bf_r = rep_branch if rep_branch else None
                df_ex, path_ex = export_to_excel(branch_filter=bf_r, date_from=exp_from_r.isoformat(), date_to=exp_to_r.isoformat())
                if df_ex.empty: st.warning("لا توجد بيانات")
                else:
                    with open(path_ex,"rb") as fh:
                        st.download_button("📥 تحميل Excel", data=fh, file_name=path_ex,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_rep")

# ══════════════════════════════════════════════════════════════════════════════
# صفحة Dashboard (الأدمن)
# ══════════════════════════════════════════════════════════════════════════════
elif st.session_state.page == "dashboard":
    if st.session_state.user_type != "admin":
        st.error("غير مصرح."); st.stop()
    try:
        import plotly.graph_objects as go_plt
        import plotly.express as px
        HAS_PLOTLY = True
    except ImportError:
        HAS_PLOTLY = False
    st.markdown("## 📊 Dashboard — نظرة عامة")
    all_vs   = fetch_visits(page=None, page_size=None)[0]
    all_fu   = fetch_follow_ups()
    all_docs = fetch_doctors()
    today_s  = date.today().isoformat()
    week_ago = (date.today() - timedelta(days=6)).isoformat()
    total_all  = len(all_vs)
    rev_all    = sum(v.get("total_price",0) for v in all_vs if v.get("status")!="ملغية")
    done_all   = sum(1 for v in all_vs if v.get("status")=="تمت")
    today_cnt  = sum(1 for v in all_vs if v.get("visit_date")==today_s)
    week_vs    = [v for v in all_vs if v.get("visit_date","")>=week_ago]
    week_rev   = sum(v.get("total_price",0) for v in week_vs if v.get("status")!="ملغية")
    unpaid_all = sum(1 for v in all_vs if v.get("payment_status","غير مدفوع")=="غير مدفوع" and v.get("status")!="ملغية")
    paid_all   = sum(1 for v in all_vs if v.get("payment_status","")=="مدفوع")
    fu_pend    = sum(1 for f in all_fu if not f.get("done"))
    canc_rate  = round(sum(1 for v in all_vs if v.get("status")=="ملغية")/max(total_all,1)*100,1)
    doc_stats  = {}
    for v in all_vs:
        doc = v.get("doctor_name","غير محدد") or "غير محدد"
        if doc not in doc_stats: doc_stats[doc] = {"count":0,"rev":0,"done":0}
        doc_stats[doc]["count"] += 1
        if v.get("status")=="تمت":   doc_stats[doc]["done"] += 1
        if v.get("status")!="ملغية": doc_stats[doc]["rev"]  += v.get("total_price",0)
    top_doc = max(doc_stats, key=lambda d: doc_stats[d]["count"]) if doc_stats else "—"
    area_counts = {}
    for v in all_vs:
        addr = v.get("address","") or ""
        part = addr.split("-")[0].strip() if "-" in addr else addr[:15].strip()
        if part: area_counts[part] = area_counts.get(part,0)+1
    top_area = max(area_counts, key=area_counts.get) if area_counts else "—"
    labs_counter = {}
    for v in all_vs:
        for line in (v.get("selected_labs_text","") or "").splitlines():
            lab = line.strip().split(" — ")[0].strip()
            if lab: labs_counter[lab] = labs_counter.get(lab,0)+1
    rated_vs  = [v for v in all_vs if v.get("rating",0)]
    avg_rating = round(sum(v["rating"] for v in rated_vs)/len(rated_vs),1) if rated_vs else 0
    branch_data = {}
    for v in all_vs:
        br = v.get("branch","غير محدد") or "غير محدد"
        if br not in branch_data: branch_data[br] = {"count":0,"rev":0,"done":0}
        branch_data[br]["count"] += 1
        if v.get("status")!="ملغية": branch_data[br]["rev"]  += v.get("total_price",0)
        if v.get("status")=="تمت":   branch_data[br]["done"] += 1
    rating_stars_kpi = "⭐" * int(avg_rating) if avg_rating else "—"
    st.markdown(f"""<div class="stat-grid">
      <div class="stat-box"><div class="stat-num">{total_all}</div><div class="stat-label">إجمالي الزيارات</div></div>
      <div class="stat-box"><div class="stat-num" style="color:#27AE60">{done_all}</div><div class="stat-label">تمت ✅</div></div>
      <div class="stat-box"><div class="stat-num">{today_cnt}</div><div class="stat-label">زيارات اليوم</div></div>
      <div class="stat-box"><div class="stat-num" style="font-size:13px">{format_money(rev_all)}</div><div class="stat-label">إجمالي الإيراد</div></div>
      <div class="stat-box"><div class="stat-num" style="font-size:13px">{format_money(week_rev)}</div><div class="stat-label">إيراد آخر 7 أيام</div></div>
    </div>
    <div class="stat-grid">
      <div class="stat-box"><div class="stat-num" style="color:#E74C3C">{unpaid_all}</div><div class="stat-label">غير مدفوع 🔴</div></div>
      <div class="stat-box"><div class="stat-num" style="color:#27AE60">{paid_all}</div><div class="stat-label">مدفوع ✅</div></div>
      <div class="stat-box"><div class="stat-num" style="color:#F39C12">{fu_pend}</div><div class="stat-label">متابعات معلقة</div></div>
      <div class="stat-box"><div class="stat-num" style="color:#E74C3C">{canc_rate}%</div><div class="stat-label">نسبة الإلغاء</div></div>
      <div class="stat-box"><div class="stat-num" style="font-size:12px;color:#9B59B6">{top_doc}</div><div class="stat-label">أكثر طبيب نشاطاً</div></div>
      <div class="stat-box"><div class="stat-num" style="font-size:11px;color:#3498DB">{top_area}</div><div class="stat-label">أكثر منطقة</div></div>
      <div class="stat-box"><div class="stat-num" style="color:#F39C12">{rating_stars_kpi}</div><div class="stat-label">متوسط التقييم {avg_rating}/5</div></div>
    </div>""", unsafe_allow_html=True)
    st.markdown("---")
    # ── بيانات شهرية ──
    monthly_d = {}
    for v in all_vs:
        if v.get("status")=="ملغية": continue
        try:
            d  = datetime.strptime(v.get("visit_date",""), "%Y-%m-%d")
            mk = f"{MONTHS_AR[d.month-1]} {d.year}"
            if mk not in monthly_d: monthly_d[mk] = {"rev":0,"count":0,"year":d.year,"month":d.month}
            monthly_d[mk]["rev"]   += v.get("total_price",0)
            monthly_d[mk]["count"] += 1
        except: pass
    sorted_m     = sorted(monthly_d.keys(), key=lambda k:(monthly_d[k]["year"],monthly_d[k]["month"]))[-6:]
    months_rev   = [monthly_d[m]["rev"]   for m in sorted_m]
    months_count = [monthly_d[m]["count"] for m in sorted_m]
    status_counts = {}
    for v in all_vs:
        s = v.get("status","مجدولة")
        status_counts[s] = status_counts.get(s,0)+1
    if HAS_PLOTLY:
        ch1, ch2 = st.columns(2)
        with ch1:
            st.markdown("##### 📅 عدد الزيارات شهرياً")
            fig = go_plt.Figure(go_plt.Bar(x=sorted_m, y=months_count, marker_color="#FF6B00",
                text=months_count, textposition="outside"))
            fig.update_layout(height=280, margin=dict(t=20,b=20,l=10,r=10),
                plot_bgcolor="#FFF8F0", paper_bgcolor="#FFF8F0", font=dict(family="Cairo"),
                yaxis=dict(showgrid=True,gridcolor="#FFE8D1"))
            st.plotly_chart(fig, use_container_width=True)
        with ch2:
            st.markdown("##### 💰 الإيراد الشهري (جنيه)")
            fig2 = go_plt.Figure(go_plt.Bar(x=sorted_m, y=months_rev, marker_color="#27AE60",
                text=[f"{r:,.0f}" for r in months_rev], textposition="outside"))
            fig2.update_layout(height=280, margin=dict(t=20,b=20,l=10,r=10),
                plot_bgcolor="#F0FFF4", paper_bgcolor="#F0FFF4", font=dict(family="Cairo"),
                yaxis=dict(showgrid=True,gridcolor="#D5F5E3"))
            st.plotly_chart(fig2, use_container_width=True)
        st.markdown("---")
        ch3, ch4 = st.columns(2)
        with ch3:
            st.markdown("##### 🔖 توزيع الحالات")
            STATUS_PIE_COLORS = {"تمت":"#27AE60","مجدولة":"#3498DB","في الطريق":"#F39C12","ملغية":"#E74C3C"}
            labels = list(status_counts.keys()); values = list(status_counts.values())
            colors = [STATUS_PIE_COLORS.get(l,"#888") for l in labels]
            fig3 = go_plt.Figure(go_plt.Pie(labels=labels, values=values,
                marker=dict(colors=colors), hole=0.4, textinfo="label+percent"))
            fig3.update_layout(height=280, margin=dict(t=20,b=20,l=10,r=10),
                font=dict(family="Cairo"), showlegend=False)
            st.plotly_chart(fig3, use_container_width=True)
        with ch4:
            st.markdown("##### 👨‍⚕️ أفضل الأطباء")
            docs_sorted_d = sorted(doc_stats.items(), key=lambda x:x[1]["count"], reverse=True)[:6]
            fig4 = go_plt.Figure(go_plt.Bar(
                y=[d[0] for d in docs_sorted_d], x=[d[1]["count"] for d in docs_sorted_d],
                orientation="h", marker_color="#9B59B6",
                text=[d[1]["count"] for d in docs_sorted_d], textposition="outside"))
            fig4.update_layout(height=280, margin=dict(t=20,b=20,l=10,r=10),
                plot_bgcolor="#F8F0FF", paper_bgcolor="#F8F0FF",
                font=dict(family="Cairo"), yaxis=dict(autorange="reversed"))
            st.plotly_chart(fig4, use_container_width=True)
        st.markdown("---")
        ch5, ch6 = st.columns(2)
        with ch5:
            st.markdown("##### 🧪 أكثر التحاليل طلباً")
            if labs_counter:
                top10 = sorted(labs_counter.items(), key=lambda x:x[1], reverse=True)[:8]
                fig5 = go_plt.Figure(go_plt.Bar(
                    y=[t[0][:20] for t in top10], x=[t[1] for t in top10],
                    orientation="h", marker_color="#E74C3C",
                    text=[t[1] for t in top10], textposition="outside"))
                fig5.update_layout(height=300, margin=dict(t=20,b=20,l=10,r=10),
                    plot_bgcolor="#FFF5F5", paper_bgcolor="#FFF5F5",
                    font=dict(family="Cairo"), yaxis=dict(autorange="reversed"))
                st.plotly_chart(fig5, use_container_width=True)
        with ch6:
            st.markdown("##### 📍 أكثر المناطق زيارةً")
            if area_counts:
                top_areas = sorted(area_counts.items(), key=lambda x:x[1], reverse=True)[:8]
                fig6 = go_plt.Figure(go_plt.Bar(
                    y=[a[0][:18] for a in top_areas], x=[a[1] for a in top_areas],
                    orientation="h", marker_color="#3498DB",
                    text=[a[1] for a in top_areas], textposition="outside"))
                fig6.update_layout(height=300, margin=dict(t=20,b=20,l=10,r=10),
                    plot_bgcolor="#EBF5FB", paper_bgcolor="#EBF5FB",
                    font=dict(family="Cairo"), yaxis=dict(autorange="reversed"))
                st.plotly_chart(fig6, use_container_width=True)
        st.markdown("---")
        st.markdown("##### 💳 توزيع حالات الدفع")
        pay_counts = {}
        for v in all_vs:
            if v.get("status")=="ملغية": continue
            ps = v.get("payment_status","غير مدفوع")
            pay_counts[ps] = pay_counts.get(ps,0)+1
        if pay_counts:
            pc_labels = list(pay_counts.keys()); pc_values = list(pay_counts.values())
            pc_colors = [PAYMENT_COLORS.get(l,"#888") for l in pc_labels]
            fig7 = go_plt.Figure(go_plt.Pie(labels=pc_labels, values=pc_values,
                marker=dict(colors=pc_colors), hole=0.4, textinfo="label+percent"))
            fig7.update_layout(height=260, margin=dict(t=10,b=10,l=10,r=10),
                font=dict(family="Cairo"), showlegend=True)
            st.plotly_chart(fig7, use_container_width=True)
        st.markdown("---")
        st.markdown("##### 🏥 مقارنة الفروع")
        if branch_data:
            df_br = pd.DataFrame([{"الفرع":b,"الزيارات":d["count"],"تمت":d["done"],"الإيراد":d["rev"]} for b,d in branch_data.items()])
            st.dataframe(df_br.style.format({"الإيراد":"{:,.0f} ج"}), use_container_width=True)
            fig8 = go_plt.Figure()
            fig8.add_trace(go_plt.Bar(name="الزيارات", x=df_br["الفرع"], y=df_br["الزيارات"], marker_color="#FF6B00"))
            fig8.add_trace(go_plt.Bar(name="تمت",       x=df_br["الفرع"], y=df_br["تمت"],       marker_color="#27AE60"))
            fig8.update_layout(height=250, barmode="group", font=dict(family="Cairo"), margin=dict(t=10,b=10,l=10,r=10))
            st.plotly_chart(fig8, use_container_width=True)
    else:
        st.warning("⚠️ أضف `plotly` في requirements.txt للرسوم البيانية المتقدمة")
        if monthly_d:
            df_m = pd.DataFrame([(m, monthly_d[m]["rev"]) for m in sorted_m], columns=["الشهر","الإيراد"])
            st.bar_chart(df_m.set_index("الشهر"))
        if doc_stats:
            df_d = pd.DataFrame([{"الدكتور":d,"الزيارات":s["count"]} for d,s in doc_stats.items()])
            st.bar_chart(df_d.set_index("الدكتور"))
    # ── تصنيفات العملاء ──
    st.markdown("---")
    st.markdown("##### 🏷️ تصنيفات العملاء")
    phones = list({v.get("phone","") for v in all_vs if v.get("phone","")})
    tag_counts = {}
    for ph in phones:
        t = get_client_tag(ph)
        tag_counts[t] = tag_counts.get(t,0)+1
    tag_grid = "".join(
        f'<div class="stat-box"><div class="stat-num" style="color:{get_client_tag_color(t)};font-size:18px">{cnt}</div>'
        f'<div class="stat-label">{t}</div></div>'
        for t,cnt in sorted(tag_counts.items(), key=lambda x:x[1], reverse=True)
    )
    st.markdown(f'<div class="stat-grid">{tag_grid}</div>', unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("### 🔒 إجراءات النظام")
    ba1, ba2 = st.columns(2)
    with ba1:
        if st.button("📥 تصدير نسخة احتياطية كاملة", use_container_width=True, key="backup_full"):
            df_bk, path_bk = export_to_excel()
            if not df_bk.empty:
                with open(path_bk,"rb") as fh:
                    st.download_button("📥 تحميل النسخة الاحتياطية", data=fh, file_name=BACKUP_EXCEL,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_backup")
    with ba2:
        uf_dash = st.file_uploader("📤 استيراد بيانات", type=["xlsx"], key="import_dash")
        if uf_dash:
            count_imported, count_updated = import_from_excel(uf_dash)
            if count_updated > 0:
                st.success(f"✅ تم استيراد {count_imported} زيارة جديدة، وتحديث {count_updated} زيارة موجودة!")
            else:
                st.success(f"✅ تم استيراد {count_imported} زيارة جديدة!")
            st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# صفحة بروفايل العميل
# ══════════════════════════════════════════════════════════════════════════════
elif st.session_state.page == "client_profile":
    phone_cp = st.session_state.get("selected_client_phone","")
    if not phone_cp:
        st.error("لم يتم تحديد عميل"); go("home")
    else:
        all_client_visits = fetch_client_history(phone_cp)
        if not all_client_visits:
            st.error("لا توجد زيارات لهذا العميل")
            if st.button("← رجوع"): go("home")
        else:
            cv = all_client_visits[0]
            tag_auto  = get_client_tag(phone_cp)
            tag_color = get_client_tag_color(tag_auto)
            churn     = get_churn_risk(phone_cp)
            total_spend = sum(v.get("total_price",0) for v in all_client_visits if v.get("status")!="ملغية")
            done_count  = sum(1 for v in all_client_visits if v.get("status")=="تمت")
            avg_spend   = round(total_spend / max(done_count,1))
            last_visit  = format_date_ar(cv.get("visit_date",""))
            unpaid_c    = sum(1 for v in all_client_visits if v.get("payment_status","غير مدفوع")=="غير مدفوع" and v.get("status")!="ملغية")
            st.markdown(f"""<div class="client-profile-header">
              <div style="display:flex;justify-content:space-between;align-items:flex-start;flex-wrap:wrap;gap:10px;">
                <div>
                  <div style="font-size:22px;font-weight:800;color:#fff;margin-bottom:4px;">👤 {cv.get('name','')}</div>
                  <div style="font-size:14px;color:rgba(255,255,255,0.8);direction:ltr;unicode-bidi:embed;">📞 {phone_cp}</div>
                  <div style="margin-top:8px;display:flex;gap:6px;flex-wrap:wrap;">
                    <span style="background:{tag_color};color:#fff;border-radius:8px;padding:3px 10px;font-size:12px;font-weight:700;">{tag_auto}</span>
                    <span style="background:rgba(255,255,255,0.15);color:#fff;border-radius:8px;padding:3px 10px;font-size:12px;">{churn}</span>
                  </div>
                </div>
                <div style="text-align:center;background:rgba(255,107,0,0.3);border-radius:12px;padding:12px 16px;">
                  <div style="font-size:28px;font-weight:800;color:#FF9A3C;">{total_spend:,}</div>
                  <div style="font-size:11px;color:rgba(255,255,255,0.7);">إجمالي الإنفاق (جنيه)</div>
                </div>
              </div>
            </div>""", unsafe_allow_html=True)
            st.markdown(f"""<div class="stat-grid">
              <div class="stat-box"><div class="stat-num">{len(all_client_visits)}</div><div class="stat-label">إجمالي الزيارات</div></div>
              <div class="stat-box"><div class="stat-num" style="color:#27AE60">{done_count}</div><div class="stat-label">زيارات مكتملة</div></div>
              <div class="stat-box"><div class="stat-num" style="font-size:16px">{avg_spend:,}</div><div class="stat-label">متوسط الزيارة</div></div>
              <div class="stat-box"><div class="stat-num" style="color:#E74C3C">{unpaid_c}</div><div class="stat-label">غير مدفوع</div></div>
            </div>""", unsafe_allow_html=True)
            pa1,pa2 = st.columns(2)
            with pa1:
                if st.button("➕ زيارة جديدة", use_container_width=True, key="cp_new"):
                    go("new", prefill={"name":cv.get("name",""),"age":cv.get("age",""),
                        "age_unit":cv.get("age_unit","سنة"),"phone":phone_cp,
                        "address":cv.get("address",""),"location_link":cv.get("location_link",""),
                        "doctor_name":cv.get("doctor_name",""),"branch":cv.get("branch","La Cite"),
                        "selected_labs_text":"","visit_time":"","notes":"",
                        "labs_price_before":0,"labs_price_after":0,"transport_fee":100})
            with pa2:
                wa_msg = f"🟠 *Orange Lab* — أهلاً بك {cv.get('name','')} 👋\nنرجو إضافة موعدك القادم معنا."
                st.markdown(f'<a href="{whatsapp_link(wa_msg, phone_cp)}" target="_blank" class="wa-btn wa-client" style="margin-top:0">📱 تواصل واتساب</a>', unsafe_allow_html=True)
            st.markdown("---")
            client_followups = fetch_follow_ups({"phone": phone_cp, "done": False})
            if client_followups:
                st.markdown('<div class="section-title">⏰ المتابعات المعلقة</div>', unsafe_allow_html=True)
                for fu in client_followups:
                    is_over = fu.get("follow_up_date","") < date.today().isoformat()
                    card_cls = "fu-card fu-overdue" if is_over else "fu-card"
                    fu_c1, fu_c2 = st.columns([5,2])
                    with fu_c1:
                        st.markdown(f'<div class="{card_cls}">{"🔴" if is_over else "⏰"} <b>{format_date_ar(fu.get("follow_up_date",""))}</b><br>{fu.get("reason","")}</div>', unsafe_allow_html=True)
                    with fu_c2:
                        if st.button("✅ تم", key=f"cp_fu_{fu["id"]}", use_container_width=True):
                            complete_follow_up(fu["id"]); st.rerun()
                st.markdown("---")
            st.markdown('<div class="section-title">📋 كل الزيارات</div>', unsafe_allow_html=True)
            for v in all_client_visits:
                vdate = format_date_ar(v.get("visit_date",""))
                vstatus = v.get("status",""); vsc = STATUS_COLORS.get(vstatus,"#888")
                vpay = v.get("payment_status","غير مدفوع"); vpc = PAYMENT_COLORS.get(vpay,"#888")
                st.markdown(f'<div class="history-card">'
                            f'<b>📅 {vdate}</b> — 👨‍⚕️ {v.get("doctor_name","")} — 🏥 {v.get("branch","")}<br>'
                            f'<span style="background:{vsc};color:#fff;border-radius:6px;padding:1px 8px;font-size:11px;">{vstatus}</span> '
                            f'<span style="background:{vpc};color:#fff;border-radius:6px;padding:1px 8px;font-size:11px;">{PAYMENT_ICONS.get(vpay,"")} {vpay}</span>'
                            f'<br>💰 {format_money(v.get("total_price",0))} — 🧪 {len((v.get("selected_labs_text","") or "").splitlines())} تحليل</div>', unsafe_allow_html=True)
                if st.button(f"📂 فتح", key=f"cp_open_{v['id']}", use_container_width=True):
                    go("detail", visit_id=v["id"])
            st.markdown("---")
            if st.button("← رجوع", use_container_width=True): go("home")

# ══════════════════════════════════════════════════════════════════════════════
# صفحة إدارة الأطباء (الأدمن فقط)
# ══════════════════════════════════════════════════════════════════════════════
elif st.session_state.page == "manage_doctors":
    if st.session_state.user_type != "admin":
        st.error("هذه الصفحة للأدمن فقط."); st.stop()
    st.markdown("### 👨‍⚕️ إدارة الأطباء")
    all_drs = fetch_doctors(active_only=False)
    active_drs   = [d for d in all_drs if d.get("active",1)]
    inactive_drs = [d for d in all_drs if not d.get("active",1)]
    st.markdown(f'<div class="stat-grid"><div class="stat-box"><div class="stat-num">{len(active_drs)}</div><div class="stat-label">نشطون</div></div><div class="stat-box"><div class="stat-num" style="color:#888">{len(inactive_drs)}</div><div class="stat-label">غير نشطين</div></div></div>', unsafe_allow_html=True)
    with st.expander("➕ إضافة دكتور جديد", expanded=False):
        nd1,nd2 = st.columns(2)
        with nd1: new_doc_name  = st.text_input("اسم الدكتور *", key="new_doc_name")
        with nd2: new_doc_phone = st.text_input("رقم التليفون", key="new_doc_phone")
        nd3,nd4,nd5 = st.columns(3)
        with nd3: new_doc_branch    = st.selectbox("الفرع", ["","La Cite","Diamond","الاثنين"], key="new_doc_branch")
        with nd4: new_doc_comm      = st.number_input("نسبة العمولة %", 0.0, 50.0, 5.0, 0.5, key="new_doc_comm")
        with nd5: new_doc_transport = st.checkbox("يحق له بدل الانتقال", key="new_doc_transport")
        if st.button("💾 حفظ الدكتور", use_container_width=True, key="save_new_doc"):
            if not new_doc_name.strip():
                st.error("أدخل اسم الدكتور")
            else:
                insert_doctor(new_doc_name.strip(), new_doc_comm, int(new_doc_transport), new_doc_phone, new_doc_branch)
                st.success(f"✅ تم إضافة الدكتور: {new_doc_name}"); st.rerun()
    st.markdown("---")
    st.markdown("#### الأطباء النشطون")
    for d in active_drs:
        doc_c1, doc_c2, doc_c3 = st.columns([4,2,2])
        with doc_c1:
            transport_badge = ' <span style="background:#27AE60;color:#fff;border-radius:6px;padding:1px 8px;font-size:10px;">🚗 بدل انتقال</span>' if d.get("transport_eligible") else ""
            st.markdown(f'<div style="padding:8px 0;"><b style="font-size:14px;">{d["name"]}</b>{transport_badge}<br>'
                        f'<span style="font-size:12px;color:#888;">عمولة {d.get("commission_pct",5)}% | {d.get("phone","") or "لا يوجد تليفون"}</span></div>',
                        unsafe_allow_html=True)
        with doc_c2:
            conn_d = get_connection()
            doc_visits = conn_d.execute(
                "SELECT COUNT(*) FROM visits WHERE doctor_name=? AND deleted_at IS NULL AND status!='ملغية'", (d["name"],)
            ).fetchone()[0]
            st.metric("زياراته", doc_visits)
        with doc_c3:
            if st.button("🚫 إيقاف", key=f"deact_{d['id']}", use_container_width=True):
                toggle_doctor_active(d["id"]); st.rerun()
    if inactive_drs:
        st.markdown("---")
        st.markdown("#### الأطباء الغير نشطين")
        for d in inactive_drs:
            dc1, dc2 = st.columns([5,2])
            with dc1:
                st.markdown(f'<div style="padding:8px 0;opacity:0.6;"><b>{d["name"]}</b> — <span style="font-size:12px;color:#888;">متوقف</span></div>', unsafe_allow_html=True)
            with dc2:
                if st.button("✅ تفعيل", key=f"act_{d['id']}", use_container_width=True):
                    toggle_doctor_active(d["id"]); st.rerun()
    st.markdown("---")
    if st.button("← رجوع للرئيسية", use_container_width=True): go("home")

# ══════════════════════════════════════════════════════════════════════════════
# صفحة المتابعات
# ══════════════════════════════════════════════════════════════════════════════
elif st.session_state.page == "follow_ups":
    if st.session_state.user_type not in ["admin","diamond","lacite"]:
        st.error("غير مصرح."); st.stop()
    st.markdown("### ⏰ المتابعات")
    today_str = date.today().isoformat()
    fu_c1, fu_c2, fu_c3 = st.columns(3)
    with fu_c1:
        fu_filter = st.selectbox("عرض", ["معلقة فقط","الكل","مكتملة"], index=0, key="fu_filter")
    with fu_c2:
        fu_date_from = st.date_input("من", value=None, key="fu_from")
    with fu_c3:
        fu_date_to   = st.date_input("إلى", value=date.today() + timedelta(days=30), key="fu_to")
    fu_filters = {}
    if fu_filter == "معلقة فقط": fu_filters["done"] = False
    elif fu_filter == "مكتملة":  fu_filters["done"] = True
    if fu_date_from: fu_filters["date_from"] = fu_date_from.isoformat()
    if fu_date_to:   fu_filters["date_to"]   = fu_date_to.isoformat()
    all_fu_list = fetch_follow_ups(fu_filters)
    overdue  = [f for f in all_fu_list if not f.get("done") and f.get("follow_up_date","") < today_str]
    due_today= [f for f in all_fu_list if not f.get("done") and f.get("follow_up_date","") == today_str]
    upcoming = [f for f in all_fu_list if not f.get("done") and f.get("follow_up_date","") > today_str]
    done_fu  = [f for f in all_fu_list if f.get("done")]
    st.markdown(f"""<div class="stat-grid">
      <div class="stat-box"><div class="stat-num" style="color:#E74C3C">{len(overdue)}</div><div class="stat-label">متأخرة 🔴</div></div>
      <div class="stat-box"><div class="stat-num" style="color:#F39C12">{len(due_today)}</div><div class="stat-label">اليوم 🟡</div></div>
      <div class="stat-box"><div class="stat-num" style="color:#3498DB">{len(upcoming)}</div><div class="stat-label">قادمة 🔵</div></div>
      <div class="stat-box"><div class="stat-num" style="color:#27AE60">{len(done_fu)}</div><div class="stat-label">مكتملة ✅</div></div>
    </div>""", unsafe_allow_html=True)
    with st.expander("➕ إضافة متابعة جديدة"):
        fnc1, fnc2 = st.columns(2)
        with fnc1: fn_name  = st.text_input("اسم العميل *", key="fn_name")
        with fnc2: fn_phone = st.text_input("التليفون", key="fn_phone")
        fnc3, fnc4 = st.columns(2)
        with fnc3: fn_date   = st.date_input("تاريخ المتابعة *", value=date.today() + timedelta(days=7), key="fn_date")
        with fnc4: fn_reason = st.text_input("السبب *", key="fn_reason")
        if st.button("💾 حفظ المتابعة", key="save_fn", use_container_width=True):
            if fn_name.strip() and fn_reason.strip():
                insert_follow_up("", fn_name.strip(), fn_phone.strip(), fn_date.isoformat(), fn_reason.strip(), created_by=st.session_state.user_email or "")
                st.success("✅ تم إضافة المتابعة!"); st.rerun()
            else: st.error("أدخل الاسم والسبب")
    st.markdown("---")
    def render_fu_list(fu_list, title, icon, card_extra_cls=""):
        if not fu_list: return
        st.markdown(f"#### {icon} {title} ({len(fu_list)})")
        for fu in fu_list:
            fu_date_disp = format_date_ar(fu.get("follow_up_date",""))
            fuc1, fuc2, fuc3 = st.columns([5,2,2])
            with fuc1:
                st.markdown(f'<div class="fu-card {card_extra_cls}">'
                            f'<b>👤 {fu.get("client_name","")}</b>'
                            f'{" — 📞 " + fu.get("client_phone","") if fu.get("client_phone") else ""}<br>'
                            f'📅 {fu_date_disp} — {fu.get("reason","")}</div>', unsafe_allow_html=True)
            with fuc2:
                if not fu.get("done"):
                    if st.button("✅ تم", key=f"fu_done_{fu['id']}", use_container_width=True):
                        complete_follow_up(fu["id"]); st.rerun()
            with fuc3:
                if fu.get("client_phone"):
                    wa_fu = f"🟠 *Orange Lab* — تذكير متابعة\n👤 {fu.get('client_name','')} — {fu.get('reason','')}\n📅 {fu_date_disp}"
                    st.markdown(f'<a href="{whatsapp_link(wa_fu, fu.get("client_phone",""))}" target="_blank" class="wa-btn wa-client" style="padding:6px 10px;font-size:11px;margin-bottom:0;">📱 واتساب</a>', unsafe_allow_html=True)
    if overdue:   render_fu_list(overdue, "متأخرة", "🔴", "fu-overdue")
    if due_today: render_fu_list(due_today, "اليوم", "🟡")
    if upcoming:  render_fu_list(upcoming, "قادمة", "🔵")
    if done_fu and fu_filter in ["الكل","مكتملة"]:
        render_fu_list(done_fu, "مكتملة", "✅", "fu-done")
    if not all_fu_list:
        st.info("لا توجد متابعات في هذه الفترة.")
    st.markdown("---")
    if st.button("← رجوع للرئيسية", use_container_width=True): go("home")

# ══════════════════════════════════════════════════════════════════════════════
# Other fallback
# ══════════════════════════════════════════════════════════════════════════════
else:
    st.warning("صفحة غير معروفة — العودة للرئيسية")
    if st.button("🏠 رجوع للرئيسية"): go("home")
